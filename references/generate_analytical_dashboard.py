import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


UNIT_ATTACH_RATE_CSV = Path("snowflake_results/unit_segment_vaps_attach_rate.csv")
MARKET_ATTACH_RATE_CSV = Path(
    "snowflake_results/unit_market_segment_vaps_attach_rate.csv"
)
DIVISION_ATTACH_RATE_CSV = Path("snowflake_results/attach_rate_unit_division.csv")
REGION_ATTACH_RATE_CSV = Path("snowflake_results/attach_rate_unit_region.csv")
RECOMMENDATION_XLSX = Path("files/actual_recomendationsheet.xlsx")
DASHBOARD_HTML = Path("files/vaps_analytical_dashboard.html")


def number(value: object) -> float:
    try:
        return float(str(value or "").replace(",", ""))
    except ValueError:
        return 0.0


def text(value: object) -> str:
    return str(value or "").strip()


def parse_vaps_header(value: object) -> tuple[str, str]:
    header = " ".join(str(value or "").replace("\r", "\n").split())
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", header)
    if not match:
        return header, ""
    return match.group(1).strip(), match.group(2).strip()


def is_covered_by_recommendation_logic(value: object) -> bool:
    value_text = str(value or "").strip()
    if not value_text:
        return False
    return value_text.lower() not in {"not applicable", "n/a"} and value_text != "0"


def recommendation_kind(value: object) -> str:
    value_text = text(value)
    lowered = value_text.lower()
    if not value_text or value_text == "0" or lowered in {"not applicable", "n/a"}:
        return "Not covered"
    if re.fullmatch(r"\d+(\.\d+)?", value_text):
        return "Fixed quantity"
    if value_text.startswith("# based on"):
        return "Quantity rule"
    if lowered.startswith("if ") or " if " in lowered:
        return "Conditional rule"
    if lowered.startswith("same as"):
        return "Dependency rule"
    return "Rule"


def load_recommendation_entries(path: Path) -> dict[tuple[str, str], dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    recommendation_entries: dict[tuple[str, str], dict[str, object]] = {}

    header_row = None
    unit_col = None
    for row in ws.iter_rows():
        for cell in row:
            if text(cell.value) == "Unit - Product Code (SF)":
                header_row = cell.row
                unit_col = cell.column
                break
        if header_row is not None:
            break

    if header_row is None or unit_col is None:
        headers = [cell.value for cell in ws[1]]
        parsed_vaps = [parse_vaps_header(header) for header in headers[7:]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            unit_code = text(row[3] if len(row) > 3 else "")
            if not unit_code:
                continue

            for index, (vaps_desc, vaps_code) in enumerate(parsed_vaps, start=7):
                if not vaps_code:
                    continue
                value = row[index] if index < len(row) else None
                value_text = text(value)
                recommendation_entries[(unit_code, vaps_code)] = {
                    "unit": unit_code,
                    "vaps": vaps_code,
                    "sequence": index - 7,
                    "vapsDesc": vaps_desc,
                    "recommendationValue": value_text,
                    "coveredByRecommendationLogic": (
                        is_covered_by_recommendation_logic(value)
                    ),
                    "recommendationKind": recommendation_kind(value),
                }
        return recommendation_entries

    vaps_columns: list[tuple[int, str, str, int]] = []
    for cell in ws[header_row]:
        vaps_desc, vaps_code = parse_vaps_header(cell.value)
        if cell.column >= 27 and re.fullmatch(r"[A-Z0-9]+", vaps_code):
            vaps_columns.append((cell.column, vaps_code, vaps_desc, len(vaps_columns)))

    for row_index in range(header_row + 1, ws.max_row + 1):
        unit_code = text(ws.cell(row_index, unit_col).value)
        if not unit_code:
            continue

        for column, vaps_code, vaps_desc, sequence in vaps_columns:
            value = ws.cell(row_index, column).value
            value_text = text(value)
            recommendation_entries[(unit_code, vaps_code)] = {
                "unit": unit_code,
                "vaps": vaps_code,
                "sequence": sequence,
                "vapsDesc": vaps_desc,
                "recommendationValue": value_text,
                "coveredByRecommendationLogic": (
                    is_covered_by_recommendation_logic(value)
                ),
                "recommendationKind": recommendation_kind(value),
            }

    return recommendation_entries


def load_rows(
    path: Path,
    segment_key: str | None,
    segment_column: str | None,
    recommendation_entries: dict[tuple[str, str], dict[str, object]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for source_row in reader:
            vaps_code = text(source_row.get("vaps_code"))
            if not vaps_code:
                continue
            unit_code = text(source_row.get("unit_code"))

            rows.append(
                {
                    "unit": unit_code,
                    "market": text(source_row.get(segment_column))
                    if segment_key == "market" and segment_column
                    else "",
                    "division": text(source_row.get("division"))
                    if segment_key == "region"
                    else text(source_row.get(segment_column))
                    if segment_key == "division" and segment_column
                    else "",
                    "region": text(source_row.get(segment_column))
                    if segment_key == "region" and segment_column
                    else "",
                    "vaps": vaps_code,
                    "coveredByRecommendationLogic": bool(
                        recommendation_entries.get((unit_code, vaps_code), {}).get(
                            "coveredByRecommendationLogic"
                        )
                    ),
                    "activations": int(number(source_row.get("Unit_Activations"))),
                    "associated": int(
                        number(source_row.get("Vaps_Associated_With_Unit"))
                    ),
                    "attachRate": number(source_row.get("Vaps_Attach_Rate")) / 100,
                    "unitName": text(source_row.get("UNIT_PRODUCTNAME_SF")),
                    "unitDescription": text(source_row.get("UNIT_DESCRIPTION")),
                    "unitL2": text(source_row.get("UNIT_L2_CORE_SOLUTION")),
                    "unitL3": text(source_row.get("UNIT_L3_PRODUCTS")),
                    "vapsDesc": text(source_row.get("VAPS_DESCRIPTION")),
                    "mainGroup": text(source_row.get("VAPS_MAIN_GROUP")) or "Unmapped",
                    "detailedGroup": text(source_row.get("VAPS_DETAILED_GROUP"))
                    or "Unmapped",
                    "tier": text(source_row.get("VAPS_PACKAGE_TIER")) or "Unmapped",
                    "source": text(source_row.get("VAPS_SOURCE")) or "Unmapped",
                }
            )
    return rows


def write_dashboard(
    unit_rows: list[dict[str, object]],
    market_rows: list[dict[str, object]],
    division_rows: list[dict[str, object]],
    region_rows: list[dict[str, object]],
    recommendation_entries: dict[tuple[str, str], dict[str, object]],
    output_path: Path,
) -> None:
    model = {
        "unitRows": unit_rows,
        "marketRows": market_rows,
        "divisionRows": division_rows,
        "regionRows": region_rows,
        "recommendationRows": sorted(
            recommendation_entries.values(),
            key=lambda row: (str(row["unit"]), str(row["vaps"])),
        ),
        "source": {
            "unitCsv": str(UNIT_ATTACH_RATE_CSV),
            "marketCsv": str(MARKET_ATTACH_RATE_CSV),
            "divisionCsv": str(DIVISION_ATTACH_RATE_CSV),
            "regionCsv": str(REGION_ATTACH_RATE_CSV),
        },
    }
    data_json = json.dumps(model, separators=(",", ":"))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Guided Selling VAPS Dashboard</title>
  <style>
    :root {{
      --ink: #17212b;
      --muted: #5f6f7d;
      --line: #d8e0e7;
      --paper: #f5f7f9;
      --panel: #ffffff;
      --green: #157c58;
      --teal: #167985;
      --blue: #2867a5;
      --red: #b64242;
      --gold: #9a6b13;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--ink);
      background: var(--paper);
      font-family: Arial, Helvetica, sans-serif;
      letter-spacing: 0;
    }}
    main {{
      display: grid;
      gap: 12px;
      max-width: 1760px;
      margin: 0 auto;
      padding: 16px;
    }}
    header, section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    header {{
      padding: 16px;
      display: grid;
      gap: 14px;
    }}
    h1 {{
      margin: 0;
      font-size: 24px;
      line-height: 1.2;
    }}
    .sub {{
      color: var(--muted);
      margin-top: 5px;
      line-height: 1.4;
    }}
    .controls {{
      display: grid;
      grid-template-columns: minmax(220px, 1.4fr) repeat(2, minmax(140px, 1fr));
      gap: 10px;
      align-items: end;
    }}
    label {{
      display: grid;
      gap: 5px;
      min-width: 0;
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
    }}
    select, input {{
      width: 100%;
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      padding: 9px 10px;
      font-size: 14px;
    }}
    .unit-card {{
      display: grid;
      grid-template-columns: minmax(240px, 1fr) repeat(3, minmax(130px, 180px));
      gap: 12px;
    }}
    .unit-info, .metric {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fbfcfd;
      padding: 12px;
      min-width: 0;
    }}
    .unit-info strong {{
      display: block;
      font-size: 17px;
      line-height: 1.25;
      overflow-wrap: anywhere;
    }}
    .unit-info span, .metric span {{
      color: var(--muted);
      display: block;
      font-size: 12px;
      margin-top: 5px;
      line-height: 1.35;
    }}
    .metric strong {{
      display: block;
      font-size: 24px;
      line-height: 1.1;
    }}
    .metric-button {{
      width: 100%;
      height: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fbfcfd;
      color: var(--ink);
      padding: 12px;
      text-align: left;
      cursor: pointer;
      font: inherit;
    }}
    .metric-button:hover {{
      border-color: var(--teal);
      background: #f3fbfb;
    }}
    .metric-button strong {{
      display: block;
      font-size: 24px;
      line-height: 1.1;
    }}
    .metric-button span {{
      color: var(--muted);
      display: block;
      font-size: 12px;
      margin-top: 5px;
      line-height: 1.35;
    }}
    .modal-backdrop {{
      position: fixed;
      inset: 0;
      z-index: 100;
      display: none;
      place-items: center;
      padding: 24px;
      background: rgba(23, 33, 43, 0.48);
    }}
    .modal-backdrop.is-open {{
      display: grid;
    }}
    .modal-panel {{
      width: min(980px, 96vw);
      max-height: 90vh;
      overflow: auto;
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 24px 70px rgba(23, 33, 43, 0.28);
    }}
    .modal-head {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: center;
      padding: 14px 16px;
      border-bottom: 1px solid var(--line);
    }}
    .modal-head h2 {{
      margin: 0;
    }}
    .modal-close {{
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      padding: 6px 10px;
      cursor: pointer;
      font-weight: 700;
    }}
    .modal-body {{
      padding: 16px;
    }}
    .elbow-chart {{
      width: 100%;
      min-height: 520px;
    }}
    .elbow-chart svg {{
      width: 100%;
      height: auto;
      display: block;
    }}
    .visual-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(320px, 1fr));
      gap: 18px;
    }}
    section {{
      padding: 12px;
      min-width: 0;
    }}
    h2 {{
      margin: 0 0 12px;
      font-size: 18px;
      line-height: 1.25;
    }}
    .section-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: baseline;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }}
    .section-head h2 {{ margin: 0; }}
    .info-tip {{
      display: inline-grid;
      place-items: center;
      width: 18px;
      height: 18px;
      margin-left: 6px;
      border: 1px solid var(--line);
      border-radius: 50%;
      color: var(--teal);
      background: #fff;
      font-size: 12px;
      font-weight: 700;
      cursor: help;
      vertical-align: middle;
      text-transform: none;
      position: relative;
    }}
    .info-tip[data-tooltip]:not(.table-info-tip)::after {{
      content: attr(data-tooltip);
      position: absolute;
      top: calc(100% + 8px);
      left: 50%;
      transform: translateX(-50%);
      z-index: 30;
      width: min(420px, 72vw);
      padding: 10px 12px;
      border-radius: 6px;
      background: var(--ink);
      color: #fff;
      box-shadow: 0 10px 24px rgba(23, 33, 43, 0.24);
      font-size: 12px;
      font-weight: 400;
      line-height: 1.45;
      text-align: left;
      text-transform: none;
      white-space: pre-line;
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
    }}
    .info-tip[data-tooltip]:not(.table-info-tip):hover::after,
    .info-tip[data-tooltip]:not(.table-info-tip):focus::after {{
      opacity: 1;
      visibility: visible;
    }}
    .section-head span {{
      color: var(--muted);
      font-size: 12px;
      text-align: right;
    }}
    .section-actions {{
      display: flex;
      align-items: center;
      gap: 10px;
      margin-left: auto;
    }}
    .download-button {{
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      padding: 7px 10px;
      font-size: 12px;
      font-weight: 700;
      cursor: pointer;
      white-space: nowrap;
    }}
    .download-button:hover {{
      border-color: var(--teal);
      color: var(--teal);
    }}
    .tooltip-panel {{
      display: none;
      margin: 0 0 8px;
      padding: 10px 12px;
      border: 1px solid #cfd9e2;
      border-radius: 6px;
      background: #f8fbfc;
      color: var(--ink);
      font-size: 12px;
      line-height: 1.45;
      white-space: pre-line;
    }}
    .tooltip-panel.is-visible {{
      display: block;
    }}
    .bars {{
      display: grid;
      gap: 8px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(165px, 1fr) minmax(160px, 1.55fr) 84px;
      gap: 10px;
      align-items: center;
    }}
    .bar-label strong, .bar-label span {{
      display: block;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .bar-label span {{
      color: var(--muted);
      font-size: 12px;
      margin-top: 2px;
    }}
    .bar-track {{
      height: 18px;
      border: 1px solid #cfd9e2;
      border-radius: 8px;
      overflow: hidden;
      background: #eef2f5;
      position: relative;
    }}
    .bar-fill {{
      height: 100%;
      background: linear-gradient(90deg, var(--teal), var(--green));
    }}
    .bar-cutoff {{
      position: absolute;
      top: 0;
      bottom: 0;
      width: 2px;
      background: var(--red);
    }}
    .bar-value {{
      text-align: right;
      font-weight: 700;
      white-space: nowrap;
    }}
    .heatmap-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      max-height: 520px;
      background: #fff;
    }}
    .heatmap {{
      width: max-content;
      min-width: 100%;
      border-collapse: collapse;
    }}
    .heatmap th, .heatmap td {{
      border-bottom: 1px solid var(--line);
      border-right: 1px solid var(--line);
      padding: 7px 8px;
      font-size: 12px;
      line-height: 1.25;
    }}
    .heatmap th {{
      position: sticky;
      top: 0;
      background: #edf2f5;
      z-index: 2;
      text-transform: uppercase;
      height: 58px;
      vertical-align: bottom;
    }}
    .heatmap th.sortable {{
      cursor: pointer;
      user-select: none;
    }}
    .heatmap th.sortable::after, .detail th::after {{
      content: " sort";
      color: var(--muted);
      font-size: 10px;
      font-weight: 700;
    }}
    .heatmap th.sort-active::after, .detail th.sort-active::after {{
      content: "  " attr(data-sort-icon);
      color: var(--teal);
    }}
    .heatmap .vaps-head {{
      width: 220px;
      min-width: 220px;
      max-width: 220px;
    }}
    .heatmap .market-head {{
      width: 122px;
      min-width: 122px;
      max-width: 122px;
      white-space: normal;
      overflow-wrap: anywhere;
      word-break: normal;
      text-align: center;
      font-size: 11px;
      line-height: 1.15;
    }}
    .market-head span {{
      display: block;
      margin-top: 4px;
      color: var(--red);
      font-weight: 700;
      text-transform: none;
    }}
    .heatmap th:first-child {{
      left: 0;
      z-index: 3;
    }}
    .heatmap td:first-child {{
      position: sticky;
      left: 0;
      z-index: 1;
      background: #fff;
      width: 220px;
      min-width: 220px;
      max-width: 220px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: normal;
    }}
    .heat-cell {{
      text-align: right;
      font-weight: 700;
      width: 122px;
      min-width: 122px;
      max-width: 122px;
      height: 58px;
      white-space: normal;
    }}
    .heat-cell span {{
      display: block;
      margin-top: 3px;
      font-size: 10px;
      font-weight: 700;
      line-height: 1.1;
    }}
    .heat-strong {{ background: #55b9a5 !important; }}
    .heat-good {{ background: #a9d9cf !important; }}
    .heat-niche {{ background: #d8edf7 !important; }}
    .heat-monitor {{ background: #f7e4b3 !important; }}
    .heat-none {{ background: #f6f8fa !important; color: #6a7682; }}
    .tables {{
      display: grid;
      gap: 18px;
    }}
    .tables section {{
      margin: 0;
    }}
    .heatmap-section {{
      min-width: 0;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      max-height: 430px;
      background: #fff;
    }}
    table.detail {{
      width: 100%;
      min-width: 1320px;
      border-collapse: collapse;
    }}
    table.resizable-columns {{
      table-layout: fixed;
      width: max-content;
      min-width: 100%;
    }}
    .detail th, .detail td {{
      border-bottom: 1px solid var(--line);
      padding: 9px 10px;
      text-align: left;
      vertical-align: top;
      font-size: 13px;
      line-height: 1.35;
    }}
    .detail th {{
      position: sticky;
      top: 0;
      background: #edf2f5;
      z-index: 1;
      text-transform: uppercase;
      font-size: 12px;
      cursor: pointer;
      user-select: none;
    }}
    .resizable-columns th {{
      padding-right: 16px;
    }}
    .resize-handle {{
      position: absolute;
      top: 0;
      right: 0;
      width: 8px;
      height: 100%;
      cursor: col-resize;
      z-index: 8;
      touch-action: none;
    }}
    .resize-handle::after {{
      content: "";
      position: absolute;
      top: 20%;
      bottom: 20%;
      right: 3px;
      width: 1px;
      background: #b9c5cf;
    }}
    body.resizing-column {{
      cursor: col-resize;
      user-select: none;
    }}
    .freeze-recommendation th:nth-child(1),
    .freeze-recommendation td:nth-child(1) {{
      position: sticky;
      left: 0;
      width: 136px;
      min-width: 136px;
      max-width: 136px;
      background: #fff;
      z-index: 2;
      box-shadow: 1px 0 0 var(--line);
    }}
    .freeze-recommendation th:nth-child(2),
    .freeze-recommendation td:nth-child(2) {{
      position: sticky;
      left: 136px;
      width: 220px;
      min-width: 220px;
      max-width: 220px;
      background: #fff;
      z-index: 2;
      box-shadow: 1px 0 0 var(--line);
    }}
    .freeze-industry th:nth-child(1),
    .freeze-industry td:nth-child(1) {{
      position: sticky;
      left: 0;
      width: var(--industry-col-1, 150px);
      min-width: var(--industry-col-1, 150px);
      max-width: var(--industry-col-1, 150px);
      background: #fff;
      z-index: 2;
      box-shadow: 1px 0 0 var(--line);
    }}
    .freeze-industry th:nth-child(2),
    .freeze-industry td:nth-child(2) {{
      position: sticky;
      left: var(--industry-col-1, 150px);
      width: var(--industry-col-2, 118px);
      min-width: var(--industry-col-2, 118px);
      max-width: var(--industry-col-2, 118px);
      background: #fff;
      z-index: 2;
      box-shadow: 1px 0 0 var(--line);
    }}
    .freeze-industry th:nth-child(3),
    .freeze-industry td:nth-child(3) {{
      position: sticky;
      left: calc(var(--industry-col-1, 150px) + var(--industry-col-2, 118px));
      width: var(--industry-col-3, 230px);
      min-width: var(--industry-col-3, 230px);
      max-width: var(--industry-col-3, 230px);
      background: #fff;
      z-index: 2;
      box-shadow: 1px 0 0 var(--line);
    }}
    .freeze-recommendation th:nth-child(-n+2),
    .freeze-industry th:nth-child(-n+3) {{
      top: 0;
      background: #edf2f5;
      z-index: 4;
    }}
    .freeze-recommendation td:nth-child(-n+2),
    .freeze-industry td:nth-child(-n+3) {{
      overflow-wrap: anywhere;
    }}
    .section-filter {{
      display: flex;
      gap: 8px;
      align-items: end;
      min-width: min(100%, 360px);
    }}
    .section-filter label {{
      width: 100%;
    }}
    .section-filters {{
      display: flex;
      gap: 10px;
      align-items: end;
      flex-wrap: wrap;
      flex: 1 1 620px;
      max-width: 760px;
    }}
    .section-filters .section-filter {{
      flex: 1 1 260px;
      min-width: 220px;
    }}
    .num {{ text-align: right !important; white-space: nowrap; }}
    .pill {{
      display: inline-block;
      border-radius: 6px;
      padding: 3px 7px;
      font-size: 12px;
      font-weight: 700;
      white-space: nowrap;
      background: #e7f4ef;
      color: var(--green);
    }}
    .pill.low {{ background: #fff2d6; color: var(--gold); }}
    .pill.none {{ background: #eef1f4; color: #52606c; }}
    .pill.add {{ background: #e8f1fb; color: var(--blue); }}
    .pill.remove {{ background: #fdecec; color: var(--red); }}
    .pill.keep {{ background: #e7f4ef; color: var(--green); }}
    .pill.promote {{ background: #e4f7f8; color: var(--teal); }}
    .pill.monitor {{ background: #fff2d6; color: var(--gold); }}
    .note {{
      color: var(--muted);
      margin: 8px 0 0;
      line-height: 1.45;
    }}
    @media (max-width: 1180px) {{
      .controls, .unit-card, .visual-grid {{ grid-template-columns: 1fr 1fr; }}
    }}
    @media (max-width: 760px) {{
      main {{ padding: 14px; }}
      .controls, .unit-card, .visual-grid {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 1fr; gap: 5px; }}
      .bar-value {{ text-align: left; }}
      h1 {{ font-size: 21px; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>Guided Selling: VAPS Recommendation for Basic GLO Units</h1>
        <div class="sub">Consolidated contract years 2024-2026. Select a unit to review VAPS attachment patterns by unit, market segment, division, and region.</div>
      </div>
      <div class="controls">
        <label>Unit
          <select id="unitFilter"></select>
        </label>
        <label>VAPS Source
          <select id="sourceFilter"></select>
        </label>
        <label>Main Group
          <select id="groupFilter"></select>
        </label>
      </div>
      <div class="unit-card" id="unitCard"></div>
    </header>

    <div class="visual-grid">
      <section>
        <div class="section-head">
          <h2>Top 5 Recommended VAPS by Attach Rate</h2>
          <span id="barSummary"></span>
        </div>
        <div class="bars" id="vapsBars"></div>
      </section>
      <section>
        <div class="section-head">
          <h2>Top 5 Missed Opportunity Rate</h2>
          <span id="missedSummary"></span>
        </div>
        <div class="bars" id="missedBars"></div>
        <p class="note">Missed opportunity means the VAPS is not covered by the actual recommendation logic and is at or above the unit elbow cutoff.</p>
      </section>
    </div>

    <div class="tables">
      <section>
        <div class="section-head">
          <h2>Recommendation Sheet Comparison</h2>
          <div class="section-filter">
            <label>Filter VAPS ID
              <input id="recommendationVapsFilter" type="search" placeholder="Example: WS1019">
            </label>
          </div>
          <div class="section-actions">
            <button class="download-button" id="downloadRecommendationCsv" type="button">Download CSV</button>
            <span id="recommendationRowCount"></span>
          </div>
        </div>
        <div class="tooltip-panel" id="recommendationTooltipPanel"></div>
        <div class="table-wrap">
          <table class="detail freeze-recommendation">
            <thead>
              <tr>
                <th data-table="recommendation" data-key="vaps">VAPS</th>
                <th data-table="recommendation" data-key="vapsDesc">VAPS description</th>
                <th data-table="recommendation" data-key="recommendationKind">Recommendation logic</th>
                <th data-table="recommendation" data-key="recommendationValue">Recommendation value</th>
                <th data-table="recommendation" data-key="coveredText">Covered</th>
                <th class="num" data-table="recommendation" data-key="activations">Unit activations</th>
                <th class="num" data-table="recommendation" data-key="associated">VAPS associated</th>
                <th class="num" data-table="recommendation" data-key="attachRate">Attach rate</th>
                <th class="num" data-table="recommendation" data-key="elbowCutoff">Unit cutoff</th>
                <th data-table="recommendation" data-key="decision">Decision <span class="info-tip table-info-tip" tabindex="0" data-tooltip-target="recommendationTooltipPanel" data-tooltip="Decision logic&#10;&#10;Keep: fixed recommendation and attach rate is at or above unit cutoff.&#10;Review Removal: fixed recommendation but attach rate is below unit cutoff.&#10;Keep Logic + Promote: conditional or quantity-driven logic is present and attach rate is at or above unit cutoff.&#10;Keep Logic: conditional or quantity-driven logic is present, but attach rate is below unit cutoff.&#10;Add: not covered in the sheet and attach rate is at or above unit cutoff.&#10;Monitor: observed attachment, but below unit cutoff.&#10;No Action: no observed attachment and not covered in the recommendation sheet.">i</span></th>
                <th data-table="recommendation" data-key="decisionReason">Reason</th>
              </tr>
            </thead>
            <tbody id="recommendationDetailBody"></tbody>
          </table>
        </div>
        <p class="note">This table starts from every VAPS column in the actual recommendation sheet for the selected unit, then joins observed 2024-2026 attach-rate metrics when available.</p>
      </section>

      <section class="heatmap-section">
        <div class="section-head">
          <h2>Market Segment Heatmap <span class="info-tip" tabindex="0" data-tooltip="Heatmap calculation&#10;&#10;Benchmark: selected unit elbow cutoff.&#10;Unit baseline: VAPS attach rate for the selected unit overall.&#10;Leverage: industry attach rate / unit attach rate.&#10;Opportunity score: max(0, industry attach rate - unit cutoff) x industry activations.&#10;Industry signal: business label based on attach rate, leverage, and opportunity score.">i</span></h2>
          <div class="section-filter">
            <label>Filter VAPS ID
              <input id="heatmapVapsFilter" type="search" placeholder="Example: WS1019">
            </label>
          </div>
          <div class="section-actions">
            <button class="download-button" id="downloadHeatmapCsv" type="button">Download CSV</button>
            <span id="heatmapSummary"></span>
          </div>
        </div>
        <div class="heatmap-wrap">
          <table class="heatmap" id="heatmapTable"></table>
        </div>
      </section>

      <section class="heatmap-section">
        <div class="section-head">
          <h2>Division Heatmap <span class="info-tip" tabindex="0" data-tooltip="Heatmap calculation&#10;&#10;Benchmark: selected unit elbow cutoff.&#10;Unit baseline: VAPS attach rate for the selected unit overall.&#10;Leverage: division attach rate / unit attach rate.&#10;Opportunity score: max(0, division attach rate - unit cutoff) x division activations.&#10;Division signal: business label based on attach rate, leverage, and opportunity score.">i</span></h2>
          <div class="section-filter">
            <label>Filter VAPS ID
              <input id="divisionHeatmapVapsFilter" type="search" placeholder="Example: WS1019">
            </label>
          </div>
          <div class="section-actions">
            <button class="download-button" id="downloadDivisionHeatmapCsv" type="button">Download CSV</button>
            <span id="divisionHeatmapSummary"></span>
          </div>
        </div>
        <div class="heatmap-wrap">
          <table class="heatmap" id="divisionHeatmapTable"></table>
        </div>
      </section>

      <section class="heatmap-section">
        <div class="section-head">
          <h2>Region Heatmap <span class="info-tip" tabindex="0" data-tooltip="Heatmap calculation&#10;&#10;Benchmark: selected unit elbow cutoff.&#10;Unit baseline: VAPS attach rate for the selected unit overall.&#10;Leverage: region attach rate / unit attach rate.&#10;Opportunity score: max(0, region attach rate - unit cutoff) x region activations.&#10;Region signal: business label based on attach rate, leverage, and opportunity score.">i</span></h2>
          <div class="section-filters">
            <div class="section-filter">
              <label>Division
                <select id="regionDivisionFilter"></select>
              </label>
            </div>
            <div class="section-filter">
              <label>Filter VAPS ID
                <input id="regionHeatmapVapsFilter" type="search" placeholder="Example: WS1019">
              </label>
            </div>
          </div>
          <div class="section-actions">
            <button class="download-button" id="downloadRegionHeatmapCsv" type="button">Download CSV</button>
            <span id="regionHeatmapSummary"></span>
          </div>
        </div>
        <div class="heatmap-wrap">
          <table class="heatmap" id="regionHeatmapTable"></table>
        </div>
      </section>

      <section>
        <div class="section-head">
          <h2>Recommendation Sheet Comparison by Industry</h2>
          <div class="section-filters">
            <div class="section-filter">
              <label>Industry
                <select id="industryComparisonFilter"></select>
              </label>
            </div>
            <div class="section-filter">
              <label>Filter VAPS ID
                <input id="industryRecommendationVapsFilter" type="search" placeholder="Example: WS1019">
              </label>
            </div>
          </div>
          <div class="section-actions">
            <button class="download-button" id="downloadIndustryRecommendationCsv" type="button">Download CSV</button>
            <span id="industryRecommendationRowCount"></span>
          </div>
        </div>
        <div class="tooltip-panel" id="industryRecommendationTooltipPanel"></div>
        <div class="table-wrap">
          <table class="detail freeze-industry resizable-columns" id="industryRecommendationTable">
            <thead>
              <tr>
                <th data-table="industryRecommendation" data-key="market">Market segment</th>
                <th data-table="industryRecommendation" data-key="vaps">VAPS</th>
                <th data-table="industryRecommendation" data-key="vapsDesc">VAPS description</th>
                <th data-table="industryRecommendation" data-key="recommendationKind">Recommendation logic</th>
                <th data-table="industryRecommendation" data-key="recommendationValue">Recommendation value</th>
                <th data-table="industryRecommendation" data-key="coveredText">Covered</th>
                <th class="num" data-table="industryRecommendation" data-key="activations">Activations</th>
                <th class="num" data-table="industryRecommendation" data-key="associated">Associated</th>
                <th class="num" data-table="industryRecommendation" data-key="attachRate">Industry attach rate</th>
                <th class="num" data-table="industryRecommendation" data-key="unitAttachRate">Unit attach rate</th>
                <th class="num" data-table="industryRecommendation" data-key="leverage">Leverage</th>
                <th class="num" data-table="industryRecommendation" data-key="opportunityScore">Opportunity score</th>
                <th data-table="industryRecommendation" data-key="industrySignal">Industry signal <span class="info-tip table-info-tip" tabindex="0" data-tooltip-target="industryRecommendationTooltipPanel" data-tooltip="Industry signal logic&#10;&#10;Benchmark: selected unit elbow cutoff.&#10;Leverage: industry attach rate / unit attach rate.&#10;Opportunity score: max(0, industry attach rate - unit cutoff) x industry activations.&#10;&#10;Strong Industry Opportunity: above benchmark and over-indexes versus the unit average.&#10;Good General Fit: above benchmark and broadly consistent with the unit average.&#10;Niche Industry Signal: strong over-indexing, but smaller opportunity volume.&#10;Monitor: observed attachment, but below benchmark.&#10;No Signal: no observed attachment in that industry.">i</span></th>
                <th data-table="industryRecommendation" data-key="industrySignalReason">Interpretation</th>
              </tr>
            </thead>
            <tbody id="industryRecommendationDetailBody"></tbody>
          </table>
        </div>
      </section>

      <section>
        <div class="section-head">
          <h2>Unit-Level VAPS Detail</h2>
          <div class="section-actions">
            <button class="download-button" id="downloadUnitCsv" type="button">Download CSV</button>
            <span id="unitRowCount"></span>
          </div>
        </div>
        <div class="table-wrap">
          <table class="detail">
            <thead>
              <tr>
                <th data-table="unit" data-key="vaps">VAPS</th>
                <th data-table="unit" data-key="vapsDesc">VAPS description</th>
                <th data-table="unit" data-key="source">Source</th>
                <th data-table="unit" data-key="mainGroup">Main group</th>
                <th data-table="unit" data-key="tier">Tier</th>
                <th class="num" data-table="unit" data-key="activations">Unit activations</th>
                <th class="num" data-table="unit" data-key="associated">VAPS associated</th>
                <th class="num" data-table="unit" data-key="attachRate">Attach rate</th>
                <th class="num" data-table="unit" data-key="elbowCutoff">Elbow cutoff</th>
                <th data-table="unit" data-key="cutoffStatus">Cutoff status</th>
              </tr>
            </thead>
            <tbody id="unitDetailBody"></tbody>
          </table>
        </div>
      </section>

      <section>
        <div class="section-head">
          <h2>Market Segment VAPS Detail</h2>
          <div class="section-actions">
            <button class="download-button" id="downloadMarketCsv" type="button">Download CSV</button>
            <span id="marketRowCount"></span>
          </div>
        </div>
        <div class="table-wrap">
          <table class="detail">
            <thead>
              <tr>
                <th data-table="market" data-key="market">Market segment</th>
                <th data-table="market" data-key="vaps">VAPS</th>
                <th data-table="market" data-key="vapsDesc">VAPS description</th>
                <th data-table="market" data-key="source">Source</th>
                <th data-table="market" data-key="mainGroup">Main group</th>
                <th data-table="market" data-key="tier">Tier</th>
                <th class="num" data-table="market" data-key="activations">Unit activations</th>
                <th class="num" data-table="market" data-key="associated">VAPS associated</th>
                <th class="num" data-table="market" data-key="attachRate">Industry attach rate</th>
                <th class="num" data-table="market" data-key="unitAttachRate">Unit attach rate</th>
                <th class="num" data-table="market" data-key="unitCutoff">Unit cutoff</th>
                <th class="num" data-table="market" data-key="leverage">Leverage</th>
                <th class="num" data-table="market" data-key="opportunityScore">Opportunity score</th>
                <th data-table="market" data-key="industrySignal">Industry signal</th>
              </tr>
            </thead>
            <tbody id="marketDetailBody"></tbody>
          </table>
        </div>
      </section>

      <section>
        <div class="section-head">
          <h2>Division VAPS Detail</h2>
          <div class="section-actions">
            <button class="download-button" id="downloadDivisionCsv" type="button">Download CSV</button>
            <span id="divisionRowCount"></span>
          </div>
        </div>
        <div class="table-wrap">
          <table class="detail">
            <thead>
              <tr>
                <th data-table="division" data-key="division">Division</th>
                <th data-table="division" data-key="vaps">VAPS</th>
                <th data-table="division" data-key="vapsDesc">VAPS description</th>
                <th data-table="division" data-key="source">Source</th>
                <th data-table="division" data-key="mainGroup">Main group</th>
                <th data-table="division" data-key="tier">Tier</th>
                <th class="num" data-table="division" data-key="activations">Unit activations</th>
                <th class="num" data-table="division" data-key="associated">VAPS associated</th>
                <th class="num" data-table="division" data-key="attachRate">Division attach rate</th>
                <th class="num" data-table="division" data-key="unitAttachRate">Unit attach rate</th>
                <th class="num" data-table="division" data-key="unitCutoff">Unit cutoff</th>
                <th class="num" data-table="division" data-key="leverage">Leverage</th>
                <th class="num" data-table="division" data-key="opportunityScore">Opportunity score</th>
                <th data-table="division" data-key="industrySignal">Division signal</th>
              </tr>
            </thead>
            <tbody id="divisionDetailBody"></tbody>
          </table>
        </div>
      </section>

      <section>
        <div class="section-head">
          <h2>Region VAPS Detail</h2>
          <div class="section-actions">
            <button class="download-button" id="downloadRegionCsv" type="button">Download CSV</button>
            <span id="regionRowCount"></span>
          </div>
        </div>
        <div class="table-wrap">
          <table class="detail">
            <thead>
              <tr>
                <th data-table="region" data-key="division">Division</th>
                <th data-table="region" data-key="region">Region</th>
                <th data-table="region" data-key="vaps">VAPS</th>
                <th data-table="region" data-key="vapsDesc">VAPS description</th>
                <th data-table="region" data-key="source">Source</th>
                <th data-table="region" data-key="mainGroup">Main group</th>
                <th data-table="region" data-key="tier">Tier</th>
                <th class="num" data-table="region" data-key="activations">Unit activations</th>
                <th class="num" data-table="region" data-key="associated">VAPS associated</th>
                <th class="num" data-table="region" data-key="attachRate">Region attach rate</th>
                <th class="num" data-table="region" data-key="unitAttachRate">Unit attach rate</th>
                <th class="num" data-table="region" data-key="unitCutoff">Unit cutoff</th>
                <th class="num" data-table="region" data-key="leverage">Leverage</th>
                <th class="num" data-table="region" data-key="opportunityScore">Opportunity score</th>
                <th data-table="region" data-key="industrySignal">Region signal</th>
              </tr>
            </thead>
            <tbody id="regionDetailBody"></tbody>
          </table>
        </div>
      </section>
    </div>
  </main>
  <div class="modal-backdrop" id="elbowModal" aria-hidden="true">
    <div class="modal-panel" role="dialog" aria-modal="true" aria-labelledby="elbowModalTitle">
      <div class="modal-head">
        <h2 id="elbowModalTitle">Elbow Chart</h2>
        <button class="modal-close" id="closeElbowModal" type="button">Close</button>
      </div>
      <div class="modal-body">
        <div class="elbow-chart" id="elbowChart"></div>
      </div>
    </div>
  </div>

  <script>
    const model = {data_json};
    const MIN_ACTIVATIONS = 25;
    const HEATMAP_VAPS_LIMIT = 50;
    const HEATMAP_MARKET_LIMIT = 10;
    const unitFilter = document.getElementById("unitFilter");
    const sourceFilter = document.getElementById("sourceFilter");
    const groupFilter = document.getElementById("groupFilter");
    const unitCard = document.getElementById("unitCard");
    const vapsBars = document.getElementById("vapsBars");
    const missedBars = document.getElementById("missedBars");
    const barSummary = document.getElementById("barSummary");
    const missedSummary = document.getElementById("missedSummary");
    const heatmapTable = document.getElementById("heatmapTable");
    const heatmapSummary = document.getElementById("heatmapSummary");
    const heatmapVapsFilter = document.getElementById("heatmapVapsFilter");
    const divisionHeatmapTable = document.getElementById("divisionHeatmapTable");
    const divisionHeatmapSummary = document.getElementById("divisionHeatmapSummary");
    const divisionHeatmapVapsFilter = document.getElementById("divisionHeatmapVapsFilter");
    const regionHeatmapTable = document.getElementById("regionHeatmapTable");
    const regionHeatmapSummary = document.getElementById("regionHeatmapSummary");
    const regionDivisionFilter = document.getElementById("regionDivisionFilter");
    const regionHeatmapVapsFilter = document.getElementById("regionHeatmapVapsFilter");
    const unitDetailBody = document.getElementById("unitDetailBody");
    const marketDetailBody = document.getElementById("marketDetailBody");
    const divisionDetailBody = document.getElementById("divisionDetailBody");
    const regionDetailBody = document.getElementById("regionDetailBody");
    const recommendationDetailBody = document.getElementById("recommendationDetailBody");
    const industryRecommendationDetailBody = document.getElementById("industryRecommendationDetailBody");
    const unitRowCount = document.getElementById("unitRowCount");
    const marketRowCount = document.getElementById("marketRowCount");
    const divisionRowCount = document.getElementById("divisionRowCount");
    const regionRowCount = document.getElementById("regionRowCount");
    const recommendationRowCount = document.getElementById("recommendationRowCount");
    const industryRecommendationRowCount = document.getElementById("industryRecommendationRowCount");
    const industryComparisonFilter = document.getElementById("industryComparisonFilter");
    const recommendationVapsFilter = document.getElementById("recommendationVapsFilter");
    const industryRecommendationVapsFilter = document.getElementById("industryRecommendationVapsFilter");
    const downloadRecommendationCsv = document.getElementById("downloadRecommendationCsv");
    const downloadIndustryRecommendationCsv = document.getElementById("downloadIndustryRecommendationCsv");
    const downloadUnitCsv = document.getElementById("downloadUnitCsv");
    const downloadMarketCsv = document.getElementById("downloadMarketCsv");
    const downloadHeatmapCsv = document.getElementById("downloadHeatmapCsv");
    const downloadDivisionCsv = document.getElementById("downloadDivisionCsv");
    const downloadDivisionHeatmapCsv = document.getElementById("downloadDivisionHeatmapCsv");
    const downloadRegionCsv = document.getElementById("downloadRegionCsv");
    const downloadRegionHeatmapCsv = document.getElementById("downloadRegionHeatmapCsv");
    const elbowModal = document.getElementById("elbowModal");
    const closeElbowModal = document.getElementById("closeElbowModal");
    const elbowChart = document.getElementById("elbowChart");
    let unitSort = {{ key: "attachRate", direction: "desc" }};
    let marketSort = {{ key: "attachRate", direction: "desc" }};
    let divisionSort = {{ key: "attachRate", direction: "desc" }};
    let regionSort = {{ key: "attachRate", direction: "desc" }};
    let recommendationSort = {{ key: "sequence", direction: "asc" }};
    let industryRecommendationSort = {{ key: "sequence", direction: "asc" }};
    let heatmapSort = {{ key: "unitAttachRate", direction: "desc" }};
    let divisionHeatmapSort = {{ key: "unitAttachRate", direction: "desc" }};
    let regionHeatmapSort = {{ key: "unitAttachRate", direction: "desc" }};
    let industryComparisonTouched = false;
    let currentRecommendationRows = [];
    let currentIndustryRecommendationRows = [];
    let currentUnitRows = [];
    let currentMarketRows = [];
    let currentDivisionRows = [];
    let currentRegionRows = [];
    let currentHeatmapRows = [];
    let currentDivisionHeatmapRows = [];
    let currentRegionHeatmapRows = [];

    function esc(value) {{
      return String(value ?? "").replace(/[&<>"']/g, char => ({{
        "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;"
      }}[char]));
    }}
    function fmtInt(value) {{ return Number(value || 0).toLocaleString(); }}
    function fmtPct(value) {{ return `${{(Number(value || 0) * 100).toFixed(1)}}%`; }}
    function fmtScore(value) {{ return Number(value || 0).toFixed(1); }}
    function fmtLeverage(value) {{
      if (value === null || value === undefined || !Number.isFinite(Number(value))) return "";
      return `${{Number(value).toFixed(2)}}x`;
    }}
    function pctValue(value) {{ return Number(value || 0) * 100; }}
    function unique(rows, key) {{
      return [...new Set(rows.map(row => row[key]).filter(Boolean))]
        .sort((a, b) => String(a).localeCompare(String(b), undefined, {{ numeric: true }}));
    }}
    function fillSelect(select, values, label) {{
      const current = select.value;
      select.innerHTML = `<option value="">${{esc(label)}}</option>` +
        values.map(value => `<option value="${{esc(value)}}">${{esc(value)}}</option>`).join("");
      if (values.includes(current)) select.value = current;
    }}
    function selectedUnitRows() {{
      const unit = unitFilter.value;
      const source = sourceFilter.value;
      const group = groupFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.unitRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount &&
        (!source || row.source === source) &&
        (!group || row.mainGroup === group)
      );
    }}
    function selectedMarketRows() {{
      const unit = unitFilter.value;
      const source = sourceFilter.value;
      const group = groupFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.marketRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount &&
        (!source || row.source === source) &&
        (!group || row.mainGroup === group)
      );
    }}
    function selectedDivisionRows() {{
      const unit = unitFilter.value;
      const source = sourceFilter.value;
      const group = groupFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.divisionRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount &&
        (!source || row.source === source) &&
        (!group || row.mainGroup === group)
      );
    }}
    function selectedRegionRows() {{
      const unit = unitFilter.value;
      const source = sourceFilter.value;
      const group = groupFilter.value;
      const division = regionDivisionFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.regionRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount &&
        (!source || row.source === source) &&
        (!group || row.mainGroup === group) &&
        (!division || row.division === division)
      );
    }}
    function allUnitRowsForCutoff() {{
      const unit = unitFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.unitRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount
      );
    }}
    function allMarketRowsForCutoff() {{
      const unit = unitFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.marketRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount
      );
    }}
    function allDivisionRowsForCutoff() {{
      const unit = unitFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.divisionRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount
      );
    }}
    function allRegionRowsForCutoff() {{
      const unit = unitFilter.value;
      const minCount = MIN_ACTIVATIONS;
      return model.regionRows.filter(row =>
        row.unit === unit &&
        row.activations >= minCount
      );
    }}
    function selectedRecommendationRows() {{
      return model.recommendationRows.filter(row => row.unit === unitFilter.value);
    }}
    function refreshIndustryComparisonOptions(marketRows) {{
      const markets = unique(marketRows, "market");
      const current = industryComparisonFilter.value;
      industryComparisonFilter.innerHTML = `<option value="">All market segments</option>` +
        markets.map(market => `<option value="${{esc(market)}}">${{esc(market)}}</option>`).join("");
      if (!current && !industryComparisonTouched && markets.length) industryComparisonFilter.value = markets[0];
      else if (!current) industryComparisonFilter.value = "";
      else if (markets.includes(current)) industryComparisonFilter.value = current;
      else if (markets.length) industryComparisonFilter.value = markets[0];
    }}
    function refreshRegionDivisionOptions() {{
      const divisions = unique(
        model.regionRows.filter(row => row.unit === unitFilter.value && row.activations >= MIN_ACTIVATIONS),
        "division"
      );
      fillSelect(regionDivisionFilter, divisions, "All divisions");
    }}
    function compareRows(key, direction) {{
      return (a, b) => {{
        const left = a[key];
        const right = b[key];
        const multiplier = direction === "asc" ? 1 : -1;
        if (typeof left === "number" && typeof right === "number") {{
          return (left - right) * multiplier;
        }}
        return String(left || "").localeCompare(String(right || ""), undefined, {{ numeric: true }}) * multiplier;
      }};
    }}
    function elbowData(rows) {{
      const rates = [...new Set(rows.filter(row => row.associated >= 25 && row.attachRate > 0).map(row => row.attachRate))]
        .sort((a, b) => b - a);
      if (!rates.length) return {{ cutoff: 0.05, cutoffIndex: -1, rates }};
      if (rates.length < 3) return {{ cutoff: rates[rates.length - 1], cutoffIndex: rates.length - 1, rates }};
      const first = rates[0];
      const last = rates[rates.length - 1];
      const denominator = Math.sqrt(Math.pow(last - first, 2) + 1);
      let bestIndex = 0;
      let bestDistance = -1;
      rates.forEach((rate, index) => {{
        const x = index / (rates.length - 1);
        const distance = Math.abs((last - first) * x - rate + first) / denominator;
        if (distance > bestDistance) {{
          bestDistance = distance;
          bestIndex = index;
        }}
      }});
      return {{ cutoff: rates[bestIndex], cutoffIndex: bestIndex, rates }};
    }}
    function elbowCutoff(rows) {{
      return elbowData(rows).cutoff;
    }}
    function decorateWithCutoff(rows, cutoff) {{
      return rows.map(row => ({{
        ...row,
        elbowCutoff: cutoff,
        cutoffStatus: row.attachRate >= cutoff ? "Above cutoff" : "Below cutoff"
      }}));
    }}
    function recommendationDecision(row, cutoff) {{
      const isFixedRecommendation = row.recommendationKind === "Fixed quantity";
      if (row.coveredByRecommendationLogic && !isFixedRecommendation && row.attachRate >= cutoff) {{
        return ["Keep Logic + Promote", `Conditional or quantity-driven logic is already present and attach rate is at or above the unit cutoff of ${{fmtPct(cutoff)}}.`];
      }}
      if (row.coveredByRecommendationLogic && !isFixedRecommendation) {{
        return ["Keep Logic", "Conditional or quantity-driven logic is already present; broad attach rate alone should not remove it."];
      }}
      if (row.coveredByRecommendationLogic && row.attachRate >= cutoff) {{
        return ["Keep", `Fixed recommendation is at or above the unit cutoff of ${{fmtPct(cutoff)}}.`];
      }}
      if (row.coveredByRecommendationLogic) {{
        return ["Review Removal", `Fixed recommendation is below the unit cutoff of ${{fmtPct(cutoff)}}.`];
      }}
      if (row.attachRate >= cutoff && row.associated > 0) {{
        return ["Add", `Not covered by recommendation logic and at or above the unit cutoff of ${{fmtPct(cutoff)}}.`];
      }}
      if (row.attachRate > 0) {{
        return ["Monitor", `Observed in transactions, but below the unit cutoff of ${{fmtPct(cutoff)}}.`];
      }}
      return ["No Action", "No observed attachment and not covered by recommendation logic."];
    }}
    function decorateRecommendationRows(recommendationRows, attachRows, cutoff) {{
      const attachByVaps = new Map(attachRows.map(row => [row.vaps, row]));
      return recommendationRows.map(row => {{
        const attach = attachByVaps.get(row.vaps) || {{}};
        const attachRate = attach.attachRate || 0;
        const associated = attach.associated || 0;
        const activations = attach.activations || 0;
        const [decision, decisionReason] = recommendationDecision({{
          ...row,
          attachRate,
          associated,
          activations
        }}, cutoff);
        return {{
          ...row,
          vapsDesc: attach.vapsDesc || row.vapsDesc || "",
          source: attach.source || "Unmapped",
          mainGroup: attach.mainGroup || "Unmapped",
          activations,
          associated,
          attachRate,
          elbowCutoff: cutoff,
          coveredText: row.coveredByRecommendationLogic ? "Yes" : "No",
          decision,
          decisionReason
        }};
      }});
    }}
    function industrySignal(row, unitCutoff) {{
      const opportunityScore = Math.max(0, row.attachRate - unitCutoff) * row.activations;
      const leverage = row.unitAttachRate > 0 ? row.attachRate / row.unitAttachRate : null;
      if (row.attachRate <= 0) {{
        return {{
          opportunityScore,
          leverage,
          industrySignal: "No Signal",
          industrySignalReason: "No observed attachment in this market segment."
        }};
      }}
      if (row.attachRate < unitCutoff) {{
        return {{
          opportunityScore,
          leverage,
          industrySignal: "Monitor",
          industrySignalReason: `Observed in this market segment, but below the unit benchmark of ${{fmtPct(unitCutoff)}}.`
        }};
      }}
      if (leverage !== null && leverage >= 1.5 && opportunityScore < 10) {{
        return {{
          opportunityScore,
          leverage,
          industrySignal: "Niche Industry Signal",
          industrySignalReason: `Over-indexes at ${{fmtLeverage(leverage)}} versus the unit average, but volume is smaller.`
        }};
      }}
      if (leverage !== null && leverage >= 1.2) {{
        return {{
          opportunityScore,
          leverage,
          industrySignal: "Strong Industry Opportunity",
          industrySignalReason: `Above the unit benchmark and over-indexes at ${{fmtLeverage(leverage)}} versus the unit average.`
        }};
      }}
      return {{
        opportunityScore,
        leverage,
        industrySignal: "Good General Fit",
        industrySignalReason: `Above the unit benchmark of ${{fmtPct(unitCutoff)}} and consistent with the unit's overall pattern.`
      }};
    }}
    function decorateIndustryRecommendationRows(recommendationRows, marketRows, selectedMarket, unitCutoff) {{
      const byVaps = new Map();
      marketRows.forEach(row => {{
        if (selectedMarket && (row.market || "Unmapped") !== selectedMarket) return;
        if (!byVaps.has(row.vaps)) byVaps.set(row.vaps, []);
        byVaps.get(row.vaps).push(row);
      }});
      const markets = selectedMarket
        ? [selectedMarket]
        : unique(marketRows, "market");

      const rows = [];
      markets.forEach(market => {{
        recommendationRows.forEach(row => {{
          const industryRow = (byVaps.get(row.vaps) || []).find(
            candidate => (candidate.market || "Unmapped") === market
          );
          const attachRate = industryRow ? industryRow.attachRate : 0;
          const associated = industryRow ? industryRow.associated : 0;
          const activations = industryRow ? industryRow.activations : 0;
          const signal = industrySignal({{
            attachRate,
            associated,
            activations,
            unitAttachRate: row.attachRate || 0
          }}, unitCutoff);
          rows.push({{
            ...row,
            market,
            vapsDesc: industryRow ? industryRow.vapsDesc || row.vapsDesc : row.vapsDesc,
            activations,
            associated,
            attachRate,
            unitAttachRate: row.attachRate || 0,
            unitCutoff,
            opportunityScore: signal.opportunityScore,
            leverage: signal.leverage,
            industrySignal: signal.industrySignal,
            industrySignalReason: signal.industrySignalReason,
            coveredText: row.coveredByRecommendationLogic ? "Yes" : "No"
          }});
        }});
      }});
      return rows;
    }}
    function decorateMarketRowsWithCutoffs(rows, cutoffBaseRows) {{
      const grouped = new Map();
      cutoffBaseRows.forEach(row => {{
        const market = row.market || "Unmapped";
        if (!grouped.has(market)) grouped.set(market, []);
        grouped.get(market).push(row);
      }});
      const cutoffs = new Map();
      grouped.forEach((marketRows, market) => {{
        cutoffs.set(market, elbowCutoff(marketRows));
      }});
      return rows.map(row => {{
        const market = row.market || "Unmapped";
        const cutoff = cutoffs.get(market) || 0.05;
        return {{
          ...row,
          marketElbowCutoff: cutoff,
          marketCutoffStatus: row.attachRate >= cutoff ? "Above industry cutoff" : "Below industry cutoff"
        }};
      }});
    }}
    function decorateSegmentRowsWithUnitBenchmark(rows, unitRows, unitCutoff) {{
      const unitAttachByVaps = new Map(unitRows.map(row => [row.vaps, row.attachRate || 0]));
      return rows.map(row => {{
        const unitAttachRate = unitAttachByVaps.get(row.vaps) || 0;
        const signal = industrySignal({{
          attachRate: row.attachRate,
          associated: row.associated,
          activations: row.activations,
          unitAttachRate
        }}, unitCutoff);
        return {{
          ...row,
          unitAttachRate,
          unitCutoff,
          opportunityScore: signal.opportunityScore,
          leverage: signal.leverage,
          industrySignal: signal.industrySignal,
          industrySignalReason: signal.industrySignalReason
        }};
      }});
    }}
    function renderUnitCard(rows, marketRows, cutoff) {{
      const base = model.unitRows.find(row => row.unit === unitFilter.value) || rows[0] || {{}};
      const activations = rows.length ? Math.max(...rows.map(row => row.activations)) : 0;
      const associated = rows.reduce((sum, row) => sum + row.associated, 0);
      unitCard.innerHTML = `
        <div class="unit-info">
          <strong>${{esc(base.unit || unitFilter.value)}}${{base.unitName ? ` | ${{esc(base.unitName)}}` : ""}}</strong>
          <span>${{esc(base.unitL2 || "")}}${{base.unitL3 ? ` | ${{esc(base.unitL3)}}` : ""}}</span>
        </div>
        <div class="metric"><strong>${{fmtInt(activations)}}</strong><span>Unit activations</span></div>
        <div class="metric"><strong>${{fmtInt(associated)}}</strong><span>VAPS associations</span></div>
        <button class="metric-button" id="openElbowChart" type="button" title="Open elbow chart">
          <strong>${{fmtPct(cutoff)}}</strong><span>Unit elbow cutoff</span>
        </button>
      `;
      document.getElementById("openElbowChart")?.addEventListener("click", openElbowModal);
    }}
    function renderTopAttachBars(rows, cutoff) {{
      const limit = 5;
      const sorted = rows
        .filter(row => row.coveredByRecommendationLogic)
        .sort((a, b) => (b.attachRate - a.attachRate) || (b.associated - a.associated));
      const topRows = sorted.slice(0, limit);
      const maxRate = Math.max(...topRows.map(row => row.attachRate), cutoff, 0.01);
      const cutoffLeft = Math.min(100, cutoff / maxRate * 100);
      barSummary.textContent = `${{topRows.length.toLocaleString()}} VAPS shown | cutoff ${{fmtPct(cutoff)}}`;
      vapsBars.innerHTML = topRows.map(row => `
        <div class="bar-row">
          <div class="bar-label">
            <strong title="${{esc(row.vapsDesc || row.vaps)}}">${{esc(row.vapsDesc || "Unmapped VAPS")}}</strong>
            <span>${{esc(row.vaps)}} | ${{fmtInt(row.associated)}} associated</span>
          </div>
          <div class="bar-track" title="Attach rate ${{fmtPct(row.attachRate)}}; cutoff ${{fmtPct(cutoff)}}">
            <div class="bar-fill" style="width:${{Math.max(2, row.attachRate / maxRate * 100)}}%"></div>
            <div class="bar-cutoff" style="left:${{cutoffLeft}}%"></div>
          </div>
          <div class="bar-value">${{fmtPct(row.attachRate)}}</div>
        </div>
      `).join("") || `<p class="note">No recommended VAPS rows match the current filters.</p>`;
    }}
    function renderMissedBars(rows, cutoff) {{
      const missedRows = rows
        .filter(row => !row.coveredByRecommendationLogic && row.attachRate >= cutoff)
        .sort((a, b) => (b.attachRate - a.attachRate) || (b.associated - a.associated))
        .slice(0, 5);
      const maxRate = Math.max(...missedRows.map(row => row.attachRate), cutoff, 0.01);
      const cutoffLeft = Math.min(100, cutoff / maxRate * 100);
      missedSummary.textContent = `${{missedRows.length.toLocaleString()}} VAPS shown | cutoff ${{fmtPct(cutoff)}}`;
      missedBars.innerHTML = missedRows.map(row => `
        <div class="bar-row">
          <div class="bar-label">
            <strong title="${{esc(row.vapsDesc || row.vaps)}}">${{esc(row.vapsDesc || "Unmapped VAPS")}}</strong>
            <span>${{esc(row.vaps)}} | ${{fmtInt(row.associated)}} associated</span>
          </div>
          <div class="bar-track" title="Missed opportunity rate ${{fmtPct(row.attachRate)}}; cutoff ${{fmtPct(cutoff)}}">
            <div class="bar-fill" style="width:${{Math.max(2, row.attachRate / maxRate * 100)}}%"></div>
            <div class="bar-cutoff" style="left:${{cutoffLeft}}%"></div>
          </div>
          <div class="bar-value">${{fmtPct(row.attachRate)}}</div>
        </div>
      `).join("") || `<p class="note">No missed opportunities above the unit cutoff match the current filters.</p>`;
    }}
    function renderElbowChart(rows) {{
      const eligible = rows
        .filter(row => row.associated >= 25 && row.attachRate > 0)
        .sort((a, b) => (b.attachRate - a.attachRate) || (b.associated - a.associated));
      const data = elbowData(rows);
      const cutoff = data.cutoff;
      const cutoffIndex = eligible.findIndex(row => row.attachRate <= cutoff);
      if (!eligible.length) {{
        elbowChart.innerHTML = `<p class="note">No VAPS with at least 25 associations are available for the selected unit.</p>`;
        return;
      }}

      const width = 920;
      const height = 520;
      const margin = {{ top: 42, right: 28, bottom: 56, left: 56 }};
      const plotWidth = width - margin.left - margin.right;
      const plotHeight = height - margin.top - margin.bottom;
      const maxRate = Math.max(...eligible.map(row => row.attachRate), cutoff, 0.01);
      const xFor = index => margin.left + (eligible.length === 1 ? 0 : index / (eligible.length - 1) * plotWidth);
      const yFor = rate => margin.top + (1 - rate / maxRate) * plotHeight;
      const points = eligible.map((row, index) => `${{xFor(index).toFixed(1)}},${{yFor(row.attachRate).toFixed(1)}}`).join(" ");
      const cutoffY = yFor(cutoff);
      const dotIndex = cutoffIndex >= 0 ? cutoffIndex : eligible.length - 1;
      const dotX = xFor(dotIndex);
      const dotY = yFor(eligible[dotIndex].attachRate);
      const selectedUnit = unitFilter.value || "Selected unit";

      elbowChart.innerHTML = `
        <svg viewBox="0 0 ${{width}} ${{height}}" role="img" aria-label="Elbow chart for ${{esc(selectedUnit)}}">
          <rect x="0" y="0" width="${{width}}" height="${{height}}" fill="#ffffff"></rect>
          <text x="${{margin.left}}" y="24" fill="#17212b" font-size="17" font-weight="700">Elbow Chart</text>
          <text x="${{width - margin.right}}" y="24" fill="#52606c" font-size="13" text-anchor="end">${{esc(selectedUnit)}} | ${{eligible.length}} VAPS with 25+ associations | cutoff ${{fmtPct(cutoff)}}</text>
          <line x1="${{margin.left}}" y1="${{margin.top}}" x2="${{margin.left}}" y2="${{height - margin.bottom}}" stroke="#d8e0e7"></line>
          <line x1="${{margin.left}}" y1="${{height - margin.bottom}}" x2="${{width - margin.right}}" y2="${{height - margin.bottom}}" stroke="#d8e0e7"></line>
          <text x="${{margin.left - 8}}" y="${{margin.top + 4}}" fill="#5f6f7d" font-size="12" text-anchor="end">${{fmtPct(maxRate)}}</text>
          <text x="${{margin.left - 8}}" y="${{height - margin.bottom + 4}}" fill="#5f6f7d" font-size="12" text-anchor="end">0.0%</text>
          <polyline points="${{points}}" fill="none" stroke="#2867a5" stroke-width="2.5"></polyline>
          <line x1="${{margin.left}}" y1="${{cutoffY.toFixed(1)}}" x2="${{width - margin.right}}" y2="${{cutoffY.toFixed(1)}}" stroke="#b64242" stroke-width="1.5" stroke-dasharray="6 6"></line>
          <circle cx="${{dotX.toFixed(1)}}" cy="${{dotY.toFixed(1)}}" r="5" fill="#b64242"></circle>
          <text x="${{margin.left + 8}}" y="${{Math.max(margin.top + 14, cutoffY - 10).toFixed(1)}}" fill="#b64242" font-size="13">Elbow cutoff: ${{fmtPct(cutoff)}}</text>
          <text x="${{margin.left + plotWidth / 2}}" y="${{height - 18}}" fill="#5f6f7d" font-size="13" text-anchor="middle">VAPS ranked by attach rate</text>
        </svg>
      `;
    }}
    function openElbowModal() {{
      renderElbowChart(allUnitRowsForCutoff());
      elbowModal.classList.add("is-open");
      elbowModal.setAttribute("aria-hidden", "false");
    }}
    function closeElbowModalPanel() {{
      elbowModal.classList.remove("is-open");
      elbowModal.setAttribute("aria-hidden", "true");
    }}
    function heatColor(rate, maxRate) {{
      if (!rate) return "#f6f8fa";
      const intensity = Math.min(1, rate / Math.max(maxRate, 0.01));
      const light = 96 - intensity * 44;
      return `hsl(170, 45%, ${{light}}%)`;
    }}
    function heatSignalClass(signal) {{
      if (signal === "Strong Industry Opportunity") return "heat-strong";
      if (signal === "Good General Fit") return "heat-good";
      if (signal === "Niche Industry Signal") return "heat-niche";
      if (signal === "Monitor") return "heat-monitor";
      return "heat-none";
    }}
    function renderSegmentHeatmap(unitRows, segmentRows, unitCutoff, options) {{
      const {{
        segmentKey,
        segmentLabel,
        segmentPluralLabel,
        signalLabel,
        filterValue,
        tableElement,
        summaryElement,
        sortState,
        currentRowsSetter
      }} = options;
      const vapsLimit = HEATMAP_VAPS_LIMIT;
      const segmentLimit = HEATMAP_MARKET_LIMIT;
      const filteredUnitRows = unitRows.filter(row => matchesVapsFilter(row, filterValue));
      const segmentTotals = new Map();
      segmentRows.forEach(row => {{
        const segment = row[segmentKey] || "Unmapped";
        segmentTotals.set(segment, Math.max(segmentTotals.get(segment) || 0, row.activations));
      }});
      const segments = [...segmentTotals.entries()]
        .sort((a, b) => b[1] - a[1])
        .slice(0, segmentLimit)
        .map(([segment]) => segment);
      const byKey = new Map(segmentRows.map(row => [`${{row.vaps}}\\u0001${{row[segmentKey] || "Unmapped"}}`, row]));
      const direction = sortState.direction === "asc" ? 1 : -1;
      const sortedUnitRows = [...filteredUnitRows].sort((a, b) => {{
        if (sortState.key === "vaps") {{
          return String(a.vaps || "").localeCompare(String(b.vaps || ""), undefined, {{ numeric: true }}) * direction;
        }}
        if (sortState.key.startsWith(`${{segmentKey}}:`)) {{
          const segment = sortState.key.slice(segmentKey.length + 1);
          const left = byKey.get(`${{a.vaps}}\\u0001${{segment}}`)?.attachRate || 0;
          const right = byKey.get(`${{b.vaps}}\\u0001${{segment}}`)?.attachRate || 0;
          return ((left - right) || ((a.attachRate || 0) - (b.attachRate || 0))) * direction;
        }}
        return (((a.attachRate || 0) - (b.attachRate || 0)) || ((a.associated || 0) - (b.associated || 0))) * direction;
      }});
      const vapsList = sortedUnitRows.slice(0, vapsLimit).map(row => row.vaps);
      const unitAttachByVaps = new Map(unitRows.map(row => [row.vaps, row.attachRate || 0]));
      currentRowsSetter(vapsList.map(vaps => {{
        const unitRow = unitRows.find(row => row.vaps === vaps) || {{}};
        const csvRow = {{
          "VAPS": vaps,
          "VAPS Description": unitRow.vapsDesc || ""
        }};
        segments.forEach(segment => {{
          const row = byKey.get(`${{vaps}}\\u0001${{segment}}`);
          const signal = row ? industrySignal({{
            attachRate: row.attachRate,
            associated: row.associated,
            activations: row.activations,
            unitAttachRate: unitAttachByVaps.get(vaps) || 0
          }}, unitCutoff) : {{
            industrySignal: "No Signal",
            opportunityScore: 0,
            leverage: null
          }};
          csvRow[`${{segment}} Attach Rate`] = row ? fmtPct(row.attachRate) : "";
          csvRow[`${{segment}} ${{signalLabel}}`] = signal.industrySignal;
          csvRow[`${{segment}} Associated`] = row ? row.associated : 0;
          csvRow[`${{segment}} Activations`] = row ? row.activations : 0;
        }});
        return csvRow;
      }}));
      summaryElement.textContent = `${{vapsList.length}} of ${{unitRows.length.toLocaleString()}} VAPS x ${{segments.length}} ${{segmentPluralLabel.toLowerCase()}} | unit cutoff ${{fmtPct(unitCutoff)}}`;
      tableElement.innerHTML = `
        <thead>
          <tr>
            <th class="vaps-head sortable ${{sortState.key === "vaps" ? "sort-active" : ""}}" data-heatmap-sort="vaps" data-sort-icon="${{sortState.direction}}">VAPS</th>
            ${{segments.map(segment => `<th class="market-head sortable ${{sortState.key === `${{segmentKey}}:${{segment}}` ? "sort-active" : ""}}" data-heatmap-sort="${{segmentKey}}:${{esc(segment)}}" data-sort-icon="${{sortState.direction}}" title="${{esc(segment)}}">${{esc(segment)}}</th>`).join("")}}
          </tr>
        </thead>
        <tbody>
          ${{vapsList.map(vaps => {{
            const unitRow = unitRows.find(row => row.vaps === vaps) || {{}};
            return `<tr>
              <td title="${{esc(unitRow.vapsDesc || vaps)}}"><strong>${{esc(vaps)}}</strong><br>${{esc(unitRow.vapsDesc || "")}}</td>
              ${{segments.map(segment => {{
                const row = byKey.get(`${{vaps}}\\u0001${{segment}}`);
                const rate = row ? row.attachRate : 0;
                const signal = row ? industrySignal({{
                  attachRate: row.attachRate,
                  associated: row.associated,
                  activations: row.activations,
                  unitAttachRate: unitAttachByVaps.get(vaps) || 0
                }}, unitCutoff) : {{
                  industrySignal: "No Signal",
                  industrySignalReason: "No observed attachment in this market segment.",
                  opportunityScore: 0,
                  leverage: null
                }};
                const unitRate = unitAttachByVaps.get(vaps) || 0;
                const title = row
                  ? `${{segment}} | ${{vaps}}
${{segmentLabel}} attach rate: ${{fmtPct(rate)}}
Unit cutoff benchmark: ${{fmtPct(unitCutoff)}}
Unit attach rate baseline: ${{fmtPct(unitRate)}}
Leverage = ${{segmentLabel.toLowerCase()}} attach rate / unit attach rate = ${{fmtLeverage(signal.leverage)}}
Opportunity score = max(0, ${{segmentLabel.toLowerCase()}} attach rate - unit cutoff) x ${{segmentLabel.toLowerCase()}} activations = ${{fmtScore(signal.opportunityScore)}}
Volume: ${{fmtInt(row.associated)}} associated / ${{fmtInt(row.activations)}} activations
${{signalLabel}}: ${{signal.industrySignal}}
Reason: ${{signal.industrySignalReason}}`
                  : `${{segment}} | ${{vaps}}
No matching ${{segmentLabel.toLowerCase()}} attachment row.
${{signalLabel}}: No Signal
Reason: No observed attachment in this ${{segmentLabel.toLowerCase()}}.`;
                return `<td class="heat-cell ${{heatSignalClass(signal.industrySignal)}}" title="${{esc(title)}}">${{row ? fmtPct(rate) : ""}}<span>${{esc(signal.industrySignal)}}</span></td>`;
              }}).join("")}}
            </tr>`;
          }}).join("")}}
        </tbody>
      `;
    }}
    function actionClass(decision) {{
      if (decision === "Keep") return "pill keep";
      if (decision === "Keep Logic + Promote") return "pill promote";
      if (decision === "Keep Logic") return "pill keep";
      if (decision === "Add") return "pill add";
      if (decision === "Review Removal") return "pill remove";
      if (decision === "Monitor") return "pill monitor";
      return "pill none";
    }}
    function signalClass(signal) {{
      if (signal === "Strong Industry Opportunity") return "pill promote";
      if (signal === "Good General Fit") return "pill keep";
      if (signal === "Niche Industry Signal") return "pill add";
      if (signal === "Monitor") return "pill monitor";
      return "pill none";
    }}
    function decisionTooltip(row) {{
      return [
        `Decision: ${{row.decision}}`,
        `Reason: ${{row.decisionReason}}`,
        `Recommendation logic: ${{row.recommendationKind || "Not covered"}}`,
        `Covered in sheet: ${{row.coveredText}}`,
        `Attach rate: ${{fmtPct(row.attachRate)}}`,
        `Unit cutoff: ${{fmtPct(row.elbowCutoff)}}`,
        `Volume: ${{fmtInt(row.associated)}} associated / ${{fmtInt(row.activations)}} activations`
      ].join("\\n");
    }}
    function industrySignalTooltip(row) {{
      return [
        `Industry signal: ${{row.industrySignal}}`,
        `Reason: ${{row.industrySignalReason}}`,
        `Industry attach rate: ${{fmtPct(row.attachRate)}}`,
        `Unit attach rate: ${{fmtPct(row.unitAttachRate)}}`,
        `Unit cutoff: ${{fmtPct(row.unitCutoff ?? row.elbowCutoff ?? 0)}}`,
        `Leverage: ${{fmtLeverage(row.leverage) || "n/a"}}`,
        `Opportunity score: ${{fmtScore(row.opportunityScore)}}`,
        `Volume: ${{fmtInt(row.associated)}} associated / ${{fmtInt(row.activations)}} activations`
      ].join("\\n");
    }}
    function segmentSignalTooltip(row, label) {{
      return [
        `${{label}} signal: ${{row.industrySignal}}`,
        `Reason: ${{row.industrySignalReason}}`,
        `${{label}} attach rate: ${{fmtPct(row.attachRate)}}`,
        `Unit attach rate: ${{fmtPct(row.unitAttachRate)}}`,
        `Unit cutoff: ${{fmtPct(row.unitCutoff ?? row.elbowCutoff ?? 0)}}`,
        `Leverage: ${{fmtLeverage(row.leverage) || "n/a"}}`,
        `Opportunity score: ${{fmtScore(row.opportunityScore)}}`,
        `Volume: ${{fmtInt(row.associated)}} associated / ${{fmtInt(row.activations)}} activations`
      ].join("\\n");
    }}
    function matchesVapsFilter(row, value) {{
      const filter = String(value || "").trim().toUpperCase();
      if (!filter) return true;
      return String(row.vaps || "").toUpperCase().includes(filter);
    }}
    function renderRecommendationTable(rows) {{
      const filtered = rows.filter(row => matchesVapsFilter(row, recommendationVapsFilter.value));
      const sorted = [...filtered].sort(compareRows(recommendationSort.key, recommendationSort.direction));
      currentRecommendationRows = sorted;
      recommendationDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.recommendationKind || "Not covered")}}</td>
          <td>${{esc(row.recommendationValue || "")}}</td>
          <td>${{esc(row.coveredText)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.elbowCutoff)}}</td>
          <td><span class="${{actionClass(row.decision)}}" title="${{esc(decisionTooltip(row))}}">${{esc(row.decision)}}</span></td>
          <td>${{esc(row.decisionReason)}}</td>
        </tr>
      `).join("");
      recommendationRowCount.textContent = `${{sorted.length.toLocaleString()}} of ${{rows.length.toLocaleString()}} recommendation-sheet VAPS`;
    }}
    function renderIndustryRecommendationTable(rows) {{
      const filtered = rows.filter(row => matchesVapsFilter(row, industryRecommendationVapsFilter.value));
      const sorted = [...filtered].sort(compareRows(industryRecommendationSort.key, industryRecommendationSort.direction));
      currentIndustryRecommendationRows = sorted;
      industryRecommendationDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td>${{esc(row.market || "Unmapped")}}</td>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.recommendationKind || "Not covered")}}</td>
          <td>${{esc(row.recommendationValue || "")}}</td>
          <td>${{esc(row.coveredText)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.unitAttachRate)}}</td>
          <td class="num">${{fmtLeverage(row.leverage)}}</td>
          <td class="num">${{fmtScore(row.opportunityScore)}}</td>
          <td><span class="${{signalClass(row.industrySignal)}}" title="${{esc(industrySignalTooltip(row))}}">${{esc(row.industrySignal)}}</span></td>
          <td>${{esc(row.industrySignalReason)}}</td>
        </tr>
      `).join("");
      const industryLabel = industryComparisonFilter.value || "all market segments";
      industryRecommendationRowCount.textContent = `${{sorted.length.toLocaleString()}} of ${{rows.length.toLocaleString()}} rows | ${{industryLabel}}`;
      syncResizableColumnWidths(document.getElementById("industryRecommendationTable"));
    }}
    function renderUnitTable(rows) {{
      const sorted = decorateWithCutoff(rows, rows[0]?.elbowCutoff || 0.05).sort(compareRows(unitSort.key, unitSort.direction));
      currentUnitRows = sorted;
      unitDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.source)}}</td>
          <td>${{esc(row.mainGroup)}}</td>
          <td>${{esc(row.tier)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.elbowCutoff)}}</td>
          <td><span class="pill ${{row.cutoffStatus === "Above cutoff" ? "" : "low"}}">${{esc(row.cutoffStatus)}}</span></td>
        </tr>
      `).join("");
      unitRowCount.textContent = `${{sorted.length.toLocaleString()}} rows`;
    }}
    function renderMarketTable(rows) {{
      const sorted = [...rows].sort(compareRows(marketSort.key, marketSort.direction));
      currentMarketRows = sorted;
      marketDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td>${{esc(row.market || "Unmapped")}}</td>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.source)}}</td>
          <td>${{esc(row.mainGroup)}}</td>
          <td>${{esc(row.tier)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.unitAttachRate)}}</td>
          <td class="num">${{fmtPct(row.unitCutoff)}}</td>
          <td class="num">${{fmtLeverage(row.leverage)}}</td>
          <td class="num">${{fmtScore(row.opportunityScore)}}</td>
          <td><span class="${{signalClass(row.industrySignal)}}" title="${{esc(segmentSignalTooltip(row, "Market segment"))}}">${{esc(row.industrySignal)}}</span></td>
        </tr>
      `).join("");
      marketRowCount.textContent = `${{sorted.length.toLocaleString()}} rows`;
    }}
    function renderDivisionTable(rows) {{
      const sorted = [...rows].sort(compareRows(divisionSort.key, divisionSort.direction));
      currentDivisionRows = sorted;
      divisionDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td>${{esc(row.division || "Unmapped")}}</td>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.source)}}</td>
          <td>${{esc(row.mainGroup)}}</td>
          <td>${{esc(row.tier)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.unitAttachRate)}}</td>
          <td class="num">${{fmtPct(row.unitCutoff)}}</td>
          <td class="num">${{fmtLeverage(row.leverage)}}</td>
          <td class="num">${{fmtScore(row.opportunityScore)}}</td>
          <td><span class="${{signalClass(row.industrySignal)}}" title="${{esc(segmentSignalTooltip(row, "Division"))}}">${{esc(row.industrySignal)}}</span></td>
        </tr>
      `).join("");
      divisionRowCount.textContent = `${{sorted.length.toLocaleString()}} rows`;
    }}
    function renderRegionTable(rows) {{
      const sorted = [...rows].sort(compareRows(regionSort.key, regionSort.direction));
      currentRegionRows = sorted;
      regionDetailBody.innerHTML = sorted.map(row => `
        <tr>
          <td>${{esc(row.division || "Unmapped")}}</td>
          <td>${{esc(row.region || "Unmapped")}}</td>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "Unmapped")}}</td>
          <td>${{esc(row.source)}}</td>
          <td>${{esc(row.mainGroup)}}</td>
          <td>${{esc(row.tier)}}</td>
          <td class="num">${{fmtInt(row.activations)}}</td>
          <td class="num">${{fmtInt(row.associated)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num">${{fmtPct(row.unitAttachRate)}}</td>
          <td class="num">${{fmtPct(row.unitCutoff)}}</td>
          <td class="num">${{fmtLeverage(row.leverage)}}</td>
          <td class="num">${{fmtScore(row.opportunityScore)}}</td>
          <td><span class="${{signalClass(row.industrySignal)}}" title="${{esc(segmentSignalTooltip(row, "Region"))}}">${{esc(row.industrySignal)}}</span></td>
        </tr>
      `).join("");
      regionRowCount.textContent = `${{sorted.length.toLocaleString()}} rows`;
    }}
    function updateSortIndicators() {{
      document.querySelectorAll("th[data-table]").forEach(header => {{
        const state = header.dataset.table === "unit"
          ? unitSort
          : header.dataset.table === "market"
            ? marketSort
            : header.dataset.table === "division"
              ? divisionSort
            : header.dataset.table === "region"
              ? regionSort
            : header.dataset.table === "industryRecommendation"
              ? industryRecommendationSort
              : recommendationSort;
        const isActive = state.key === header.dataset.key;
        header.classList.toggle("sort-active", isActive);
        if (isActive) header.dataset.sortIcon = state.direction;
        else delete header.dataset.sortIcon;
      }});
    }}
    function refreshFilterOptions() {{
      const units = unique(model.unitRows, "unit");
      unitFilter.innerHTML = units.map(unit => `<option value="${{esc(unit)}}">${{esc(unit)}}</option>`).join("");
      unitFilter.value = units.includes("10W") ? "10W" : units[0];
      fillSelect(sourceFilter, unique(model.unitRows, "source"), "All sources");
      fillSelect(groupFilter, unique(model.unitRows, "mainGroup"), "All groups");
      refreshRegionDivisionOptions();
    }}
    function render() {{
      const rows = selectedUnitRows();
      const cutoff = elbowCutoff(allUnitRowsForCutoff());
      const marketRows = decorateMarketRowsWithCutoffs(selectedMarketRows(), allMarketRowsForCutoff());
      const benchmarkedMarketRows = decorateSegmentRowsWithUnitBenchmark(marketRows, allUnitRowsForCutoff(), cutoff);
      const divisionRows = selectedDivisionRows();
      const benchmarkedDivisionRows = decorateSegmentRowsWithUnitBenchmark(divisionRows, allUnitRowsForCutoff(), cutoff);
      refreshRegionDivisionOptions();
      const regionRows = selectedRegionRows();
      const benchmarkedRegionRows = decorateSegmentRowsWithUnitBenchmark(regionRows, allUnitRowsForCutoff(), cutoff);
      refreshIndustryComparisonOptions(marketRows);
      const decoratedRows = decorateWithCutoff(rows, cutoff);
      const recommendationRows = decorateRecommendationRows(
        selectedRecommendationRows(),
        allUnitRowsForCutoff(),
        cutoff
      );
      const industryRecommendationRows = decorateIndustryRecommendationRows(
        recommendationRows,
        marketRows,
        industryComparisonFilter.value,
        cutoff
      );
      renderUnitCard(rows, marketRows, cutoff);
      renderTopAttachBars(rows, cutoff);
      renderMissedBars(rows, cutoff);
      renderSegmentHeatmap(rows, marketRows, cutoff, {{
        segmentKey: "market",
        segmentLabel: "Market segment",
        segmentPluralLabel: "market segments",
        signalLabel: "Industry Signal",
        filterValue: heatmapVapsFilter.value,
        tableElement: heatmapTable,
        summaryElement: heatmapSummary,
        sortState: heatmapSort,
        currentRowsSetter: rows => {{
          currentHeatmapRows = rows;
        }}
      }});
      renderSegmentHeatmap(rows, divisionRows, cutoff, {{
        segmentKey: "division",
        segmentLabel: "Division",
        segmentPluralLabel: "divisions",
        signalLabel: "Division Signal",
        filterValue: divisionHeatmapVapsFilter.value,
        tableElement: divisionHeatmapTable,
        summaryElement: divisionHeatmapSummary,
        sortState: divisionHeatmapSort,
        currentRowsSetter: rows => {{
          currentDivisionHeatmapRows = rows;
        }}
      }});
      renderSegmentHeatmap(rows, regionRows, cutoff, {{
        segmentKey: "region",
        segmentLabel: "Region",
        segmentPluralLabel: "regions",
        signalLabel: "Region Signal",
        filterValue: regionHeatmapVapsFilter.value,
        tableElement: regionHeatmapTable,
        summaryElement: regionHeatmapSummary,
        sortState: regionHeatmapSort,
        currentRowsSetter: rows => {{
          currentRegionHeatmapRows = rows;
        }}
      }});
      renderRecommendationTable(recommendationRows);
      renderIndustryRecommendationTable(industryRecommendationRows);
      renderUnitTable(decoratedRows);
      renderMarketTable(benchmarkedMarketRows);
      renderDivisionTable(benchmarkedDivisionRows);
      renderRegionTable(benchmarkedRegionRows);
      updateSortIndicators();
    }}
    function csvValue(value) {{
      const text = String(value ?? "");
      return /[",\\r\\n]/.test(text) ? `"${{text.replace(/"/g, '""')}}"` : text;
    }}
    function downloadCsv(filename, columns, rows) {{
      const lines = [
        columns.map(column => csvValue(column.label)).join(","),
        ...rows.map(row => columns.map(column => csvValue(column.value(row))).join(","))
      ];
      const blob = new Blob([lines.join("\\r\\n")], {{ type: "text/csv;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = filename;
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    }}
    function fileUnit() {{
      return String(unitFilter.value || "unit").replace(/[^A-Za-z0-9_-]+/g, "_");
    }}
    downloadRecommendationCsv.addEventListener("click", () => downloadCsv(
      `recommendation_sheet_comparison_${{fileUnit()}}.csv`,
      [
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Recommendation Logic", value: row => row.recommendationKind || "Not covered" }},
        {{ label: "Recommendation Value", value: row => row.recommendationValue || "" }},
        {{ label: "Covered", value: row => row.coveredText }},
        {{ label: "Unit Activations", value: row => row.activations }},
        {{ label: "VAPS Associated", value: row => row.associated }},
        {{ label: "Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Unit Cutoff", value: row => fmtPct(row.elbowCutoff) }},
        {{ label: "Decision", value: row => row.decision }},
        {{ label: "Reason", value: row => row.decisionReason }}
      ],
      currentRecommendationRows
    ));
    downloadIndustryRecommendationCsv.addEventListener("click", () => downloadCsv(
      `recommendation_sheet_comparison_by_industry_${{fileUnit()}}.csv`,
      [
        {{ label: "Market Segment", value: row => row.market || "Unmapped" }},
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Recommendation Logic", value: row => row.recommendationKind || "Not covered" }},
        {{ label: "Recommendation Value", value: row => row.recommendationValue || "" }},
        {{ label: "Covered", value: row => row.coveredText }},
        {{ label: "Activations", value: row => row.activations }},
        {{ label: "Associated", value: row => row.associated }},
        {{ label: "Industry Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Unit Attach Rate", value: row => fmtPct(row.unitAttachRate) }},
        {{ label: "Leverage", value: row => fmtLeverage(row.leverage) }},
        {{ label: "Opportunity Score", value: row => fmtScore(row.opportunityScore) }},
        {{ label: "Industry Signal", value: row => row.industrySignal }},
        {{ label: "Interpretation", value: row => row.industrySignalReason }}
      ],
      currentIndustryRecommendationRows
    ));
    downloadUnitCsv.addEventListener("click", () => downloadCsv(
      `unit_level_vaps_detail_${{fileUnit()}}.csv`,
      [
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Source", value: row => row.source }},
        {{ label: "Main Group", value: row => row.mainGroup }},
        {{ label: "Tier", value: row => row.tier }},
        {{ label: "Unit Activations", value: row => row.activations }},
        {{ label: "VAPS Associated", value: row => row.associated }},
        {{ label: "Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Elbow Cutoff", value: row => fmtPct(row.elbowCutoff) }},
        {{ label: "Cutoff Status", value: row => row.cutoffStatus }}
      ],
      currentUnitRows
    ));
    downloadMarketCsv.addEventListener("click", () => downloadCsv(
      `market_segment_vaps_detail_${{fileUnit()}}.csv`,
      [
        {{ label: "Market Segment", value: row => row.market || "Unmapped" }},
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Source", value: row => row.source }},
        {{ label: "Main Group", value: row => row.mainGroup }},
        {{ label: "Tier", value: row => row.tier }},
        {{ label: "Activations", value: row => row.activations }},
        {{ label: "Associated", value: row => row.associated }},
        {{ label: "Industry Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Unit Attach Rate", value: row => fmtPct(row.unitAttachRate) }},
        {{ label: "Unit Cutoff", value: row => fmtPct(row.unitCutoff) }},
        {{ label: "Leverage", value: row => fmtLeverage(row.leverage) }},
        {{ label: "Opportunity Score", value: row => fmtScore(row.opportunityScore) }},
        {{ label: "Industry Signal", value: row => row.industrySignal }}
      ],
      currentMarketRows
    ));
    downloadDivisionCsv.addEventListener("click", () => downloadCsv(
      `division_vaps_detail_${{fileUnit()}}.csv`,
      [
        {{ label: "Division", value: row => row.division || "Unmapped" }},
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Source", value: row => row.source }},
        {{ label: "Main Group", value: row => row.mainGroup }},
        {{ label: "Tier", value: row => row.tier }},
        {{ label: "Activations", value: row => row.activations }},
        {{ label: "Associated", value: row => row.associated }},
        {{ label: "Division Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Unit Attach Rate", value: row => fmtPct(row.unitAttachRate) }},
        {{ label: "Unit Cutoff", value: row => fmtPct(row.unitCutoff) }},
        {{ label: "Leverage", value: row => fmtLeverage(row.leverage) }},
        {{ label: "Opportunity Score", value: row => fmtScore(row.opportunityScore) }},
        {{ label: "Division Signal", value: row => row.industrySignal }}
      ],
      currentDivisionRows
    ));
    downloadRegionCsv.addEventListener("click", () => downloadCsv(
      `region_vaps_detail_${{fileUnit()}}.csv`,
      [
        {{ label: "Division", value: row => row.division || "Unmapped" }},
        {{ label: "Region", value: row => row.region || "Unmapped" }},
        {{ label: "VAPS", value: row => row.vaps }},
        {{ label: "VAPS Description", value: row => row.vapsDesc || "Unmapped" }},
        {{ label: "Source", value: row => row.source }},
        {{ label: "Main Group", value: row => row.mainGroup }},
        {{ label: "Tier", value: row => row.tier }},
        {{ label: "Activations", value: row => row.activations }},
        {{ label: "Associated", value: row => row.associated }},
        {{ label: "Region Attach Rate", value: row => fmtPct(row.attachRate) }},
        {{ label: "Unit Attach Rate", value: row => fmtPct(row.unitAttachRate) }},
        {{ label: "Unit Cutoff", value: row => fmtPct(row.unitCutoff) }},
        {{ label: "Leverage", value: row => fmtLeverage(row.leverage) }},
        {{ label: "Opportunity Score", value: row => fmtScore(row.opportunityScore) }},
        {{ label: "Region Signal", value: row => row.industrySignal }}
      ],
      currentRegionRows
    ));
    downloadHeatmapCsv.addEventListener("click", () => {{
      const labels = currentHeatmapRows.length ? Object.keys(currentHeatmapRows[0]) : ["VAPS", "VAPS Description"];
      downloadCsv(
        `market_segment_heatmap_${{fileUnit()}}.csv`,
        labels.map(label => ({{ label, value: row => row[label] }})),
        currentHeatmapRows
      );
    }});
    downloadDivisionHeatmapCsv.addEventListener("click", () => {{
      const labels = currentDivisionHeatmapRows.length ? Object.keys(currentDivisionHeatmapRows[0]) : ["VAPS", "VAPS Description"];
      downloadCsv(
        `division_heatmap_${{fileUnit()}}.csv`,
        labels.map(label => ({{ label, value: row => row[label] }})),
        currentDivisionHeatmapRows
      );
    }});
    downloadRegionHeatmapCsv.addEventListener("click", () => {{
      const labels = currentRegionHeatmapRows.length ? Object.keys(currentRegionHeatmapRows[0]) : ["VAPS", "VAPS Description"];
      downloadCsv(
        `region_heatmap_${{fileUnit()}}.csv`,
        labels.map(label => ({{ label, value: row => row[label] }})),
        currentRegionHeatmapRows
      );
    }});
    function hideTableTooltipPanels() {{
      document.querySelectorAll(".tooltip-panel").forEach(panel => {{
        panel.classList.remove("is-visible");
        panel.textContent = "";
      }});
    }}
    function showTableTooltip(trigger) {{
      const panel = document.getElementById(trigger.dataset.tooltipTarget);
      if (!panel) return;
      hideTableTooltipPanels();
      panel.textContent = trigger.dataset.tooltip || "";
      panel.classList.add("is-visible");
    }}
    document.querySelectorAll(".table-info-tip[data-tooltip-target]").forEach(trigger => {{
      trigger.addEventListener("mouseenter", () => showTableTooltip(trigger));
      trigger.addEventListener("focus", () => showTableTooltip(trigger));
      trigger.addEventListener("click", event => {{
        event.stopPropagation();
        showTableTooltip(trigger);
      }});
    }});
    document.addEventListener("click", event => {{
      if (!event.target.closest(".table-info-tip")) hideTableTooltipPanels();
    }});
    closeElbowModal.addEventListener("click", closeElbowModalPanel);
    elbowModal.addEventListener("click", event => {{
      if (event.target === elbowModal) closeElbowModalPanel();
    }});
    document.addEventListener("keydown", event => {{
      if (event.key === "Escape" && elbowModal.classList.contains("is-open")) closeElbowModalPanel();
    }});
    function setTableColumnWidth(table, columnIndex, width) {{
      const safeWidth = Math.max(70, Math.round(width));
      table.querySelectorAll("tr").forEach(row => {{
        const cell = row.children[columnIndex];
        if (!cell) return;
        cell.style.width = `${{safeWidth}}px`;
        cell.style.minWidth = `${{safeWidth}}px`;
        cell.style.maxWidth = `${{safeWidth}}px`;
      }});
      if (table.id === "industryRecommendationTable" && columnIndex < 3) {{
        table.style.setProperty(`--industry-col-${{columnIndex + 1}}`, `${{safeWidth}}px`);
      }}
    }}
    function syncResizableColumnWidths(table) {{
      if (!table) return;
      [...table.querySelectorAll("thead th")].forEach((header, columnIndex) => {{
        const width = Number.parseFloat(header.style.width) || header.offsetWidth;
        setTableColumnWidth(table, columnIndex, width);
      }});
    }}
    function initializeResizableColumns(table) {{
      if (!table || table.dataset.resizableReady === "true") return;
      table.dataset.resizableReady = "true";
      const headers = [...table.querySelectorAll("thead th")];
      headers.forEach((header, columnIndex) => {{
        setTableColumnWidth(table, columnIndex, header.offsetWidth);
        const handle = document.createElement("span");
        handle.className = "resize-handle";
        handle.title = "Drag to resize column";
        header.appendChild(handle);

        handle.addEventListener("click", event => event.stopPropagation());
        handle.addEventListener("mousedown", event => {{
          event.preventDefault();
          event.stopPropagation();
          const startX = event.clientX;
          const startWidth = header.offsetWidth;
          document.body.classList.add("resizing-column");

          const onMove = moveEvent => {{
            setTableColumnWidth(table, columnIndex, startWidth + moveEvent.clientX - startX);
          }};
          const onUp = () => {{
            document.body.classList.remove("resizing-column");
            document.removeEventListener("mousemove", onMove);
            document.removeEventListener("mouseup", onUp);
          }};
          document.addEventListener("mousemove", onMove);
          document.addEventListener("mouseup", onUp);
        }});
      }});
    }}
    document.querySelectorAll("th[data-table]").forEach(header => {{
      header.addEventListener("click", () => {{
        const state = header.dataset.table === "unit"
          ? unitSort
          : header.dataset.table === "market"
            ? marketSort
            : header.dataset.table === "division"
              ? divisionSort
            : header.dataset.table === "region"
              ? regionSort
            : header.dataset.table === "industryRecommendation"
              ? industryRecommendationSort
              : recommendationSort;
        const key = header.dataset.key;
        if (state.key === key) state.direction = state.direction === "asc" ? "desc" : "asc";
        else {{
          state.key = key;
          state.direction = header.classList.contains("num") ? "desc" : "asc";
        }}
        render();
      }});
    }});
    heatmapTable.addEventListener("click", event => {{
      const header = event.target.closest("th[data-heatmap-sort]");
      if (!header) return;
      const key = header.dataset.heatmapSort;
      if (heatmapSort.key === key) heatmapSort.direction = heatmapSort.direction === "asc" ? "desc" : "asc";
      else {{
        heatmapSort.key = key;
        heatmapSort.direction = key === "vaps" ? "asc" : "desc";
      }}
      render();
    }});
    divisionHeatmapTable.addEventListener("click", event => {{
      const header = event.target.closest("th[data-heatmap-sort]");
      if (!header) return;
      const key = header.dataset.heatmapSort;
      if (divisionHeatmapSort.key === key) divisionHeatmapSort.direction = divisionHeatmapSort.direction === "asc" ? "desc" : "asc";
      else {{
        divisionHeatmapSort.key = key;
        divisionHeatmapSort.direction = key === "vaps" ? "asc" : "desc";
      }}
      render();
    }});
    regionHeatmapTable.addEventListener("click", event => {{
      const header = event.target.closest("th[data-heatmap-sort]");
      if (!header) return;
      const key = header.dataset.heatmapSort;
      if (regionHeatmapSort.key === key) regionHeatmapSort.direction = regionHeatmapSort.direction === "asc" ? "desc" : "asc";
      else {{
        regionHeatmapSort.key = key;
        regionHeatmapSort.direction = key === "vaps" ? "asc" : "desc";
      }}
      render();
    }});
    initializeResizableColumns(document.getElementById("industryRecommendationTable"));
    [unitFilter, sourceFilter, groupFilter, heatmapVapsFilter, divisionHeatmapVapsFilter, regionDivisionFilter, regionHeatmapVapsFilter, recommendationVapsFilter, industryRecommendationVapsFilter].forEach(control => {{
      control.addEventListener("input", render);
      control.addEventListener("change", render);
    }});
    industryComparisonFilter.addEventListener("change", () => {{
      industryComparisonTouched = true;
      render();
    }});
    refreshFilterOptions();
    render();
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )


def main() -> None:
    recommendation_entries = load_recommendation_entries(RECOMMENDATION_XLSX)
    unit_rows = load_rows(
        UNIT_ATTACH_RATE_CSV,
        segment_key=None,
        segment_column=None,
        recommendation_entries=recommendation_entries,
    )
    market_rows = load_rows(
        MARKET_ATTACH_RATE_CSV,
        segment_key="market",
        segment_column="MARKET_SEGMENT_DESCRIPTION",
        recommendation_entries=recommendation_entries,
    )
    division_rows = load_rows(
        DIVISION_ATTACH_RATE_CSV,
        segment_key="division",
        segment_column="division",
        recommendation_entries=recommendation_entries,
    )
    region_rows = load_rows(
        REGION_ATTACH_RATE_CSV,
        segment_key="region",
        segment_column="region",
        recommendation_entries=recommendation_entries,
    )
    write_dashboard(
        unit_rows,
        market_rows,
        division_rows,
        region_rows,
        recommendation_entries,
        DASHBOARD_HTML,
    )
    print(f"Wrote {DASHBOARD_HTML}")
    print(f"Unit rows: {len(unit_rows):,}")
    print(f"Unit + market rows: {len(market_rows):,}")
    print(f"Unit + division rows: {len(division_rows):,}")
    print(f"Unit + region rows: {len(region_rows):,}")
    print(f"Recommendation coverage keys: {len(recommendation_entries):,}")


if __name__ == "__main__":
    main()
