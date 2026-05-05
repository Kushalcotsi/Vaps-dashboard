import csv
import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ATTACH_RATE_CSV = Path("snowflake_results/unit_segment_vaps_attach_rate.csv")
MARKET_ATTACH_RATE_CSV = Path(
    "snowflake_results/unit_market_segment_vaps_attach_rate.csv"
)
RECOMMENDATION_XLSX = Path("files/recomendationsheet.xlsx")
VAPS_MASTER_CSV = Path("files/vaps_master.csv")
REPORT_HTML = Path("files/vaps_attach_rate_report.html")
MARKET_REPORT_HTML = Path("files/vaps_attach_rate_market_segment_report.html")

DEFAULT_ATTACH_RATE_CUTOFF = 0.05
MIN_MISSED_VAPS_BASKETS = 25


def suggested_action(
    recommended: bool,
    attach_rate: float,
    vaps_baskets: int,
    cutoff: float,
) -> tuple[str, str]:
    cutoff_text = pct(cutoff)
    if recommended and attach_rate >= cutoff:
        return "Keep", f"Recommended and at or above the unit elbow cutoff of {cutoff_text}."
    if recommended:
        return "Remove", f"Recommended, but below the unit elbow cutoff of {cutoff_text}."
    if attach_rate >= cutoff and vaps_baskets >= MIN_MISSED_VAPS_BASKETS:
        return "Add", f"Not recommended today and at or above the unit elbow cutoff of {cutoff_text}."
    if attach_rate > 0:
        return "Monitor", f"Not recommended today and below the unit elbow cutoff of {cutoff_text}."
    return "No Action", f"No observed attachment and below the unit elbow cutoff of {cutoff_text}."


def status_for_row(
    recommended: bool,
    attach_rate: float,
    vaps_baskets: int,
    cutoff: float,
) -> str:
    if recommended and attach_rate >= cutoff:
        return "Recommended Meets Cutoff"
    if recommended:
        return "Recommended Below Cutoff"
    if attach_rate >= cutoff and vaps_baskets >= MIN_MISSED_VAPS_BASKETS:
        return "Missed Opportunity"
    return "Below Cutoff"


def parse_vaps_header(value: object) -> tuple[str, str]:
    text = " ".join(str(value or "").replace("\r", "\n").split())
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", text)
    if not match:
        return text, ""
    return match.group(1).strip(), match.group(2).strip()


def is_recommended(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return text.lower() != "not applicable" and text != "0"


def pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def whole(value: object) -> str:
    try:
        return f"{int(float(value)):,}"
    except (TypeError, ValueError):
        return "0"


def number(value: object) -> float:
    try:
        return float(str(value or "").replace(",", ""))
    except ValueError:
        return 0.0


def integer(value: object) -> int:
    return int(number(value))


def first_value(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def load_recommendations(
    path: Path,
) -> tuple[dict[tuple[str, str], dict[str, str]], dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    vaps_headers = headers[7:]
    vaps_lookup: dict[str, str] = {}
    parsed_headers = []

    for header in vaps_headers:
        desc, vaps_id = parse_vaps_header(header)
        parsed_headers.append((desc, vaps_id))
        if vaps_id:
            vaps_lookup.setdefault(vaps_id, desc)

    recommendations: dict[tuple[str, str], dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit_code = str(row[3] or "").strip()
        if not unit_code:
            continue

        for index, (vaps_desc, vaps_id) in enumerate(parsed_headers, start=7):
            if not vaps_id:
                continue

            value = row[index] if index < len(row) else None
            recommendations[(unit_code, vaps_id)] = {
                "in_recommendation_sheet": True,
                "recommended": is_recommended(value),
                "recommendation_value": "" if value is None else str(value),
                "vaps_desc": vaps_desc,
                "unit_name": "" if row[4] is None else str(row[4]),
                "unit_description": "" if row[5] is None else str(row[5]),
                "unit_detailed_description": "" if row[6] is None else str(row[6]),
                "unit_l1_purpose": "" if row[0] is None else str(row[0]),
                "unit_l2_core_solution": "" if row[1] is None else str(row[1]),
                "unit_l3_products": "" if row[2] is None else str(row[2]),
            }

    return recommendations, vaps_lookup


def load_vaps_master(path: Path) -> dict[str, dict[str, str]]:
    master: dict[str, dict[str, str]] = {}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            vaps_id = (row.get("VAPS_ID") or "").strip()
            if not vaps_id:
                continue
            master[vaps_id] = {
                "description": (row.get("VAPS_DESCRIPTION") or "").strip(),
                "l1_purpose": (row.get("VAPS_L1_PURPOSE") or "").strip(),
                "l2_core_need": (row.get("VAPS_L2__CORE_NEED") or "").strip(),
                "l2_products": (row.get("VAPS_L2_PRODUCTS") or "").strip(),
                "customization": (
                    row.get("VAPS_CUSTOMIZATION_COMPLEMENTARY_ALTERNATIVES") or ""
                ).strip(),
                "main_group": (row.get("VAPS_MAIN_GROUP") or "").strip(),
                "detailed_group": (row.get("VAPS_DETAILED_GROUP") or "").strip(),
                "package_tier": (row.get("VAPS_PACKAGE_TIER") or "").strip(),
                "source": (row.get("VAPS_SOURCE") or "").strip(),
            }
    return master


def load_attach_rows(
    attach_path: Path,
    recommendations: dict[tuple[str, str], dict[str, str]],
    vaps_lookup: dict[str, str],
    vaps_master: dict[str, dict[str, str]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    seen_keys: set[tuple[str, str]] = set()
    unit_basket_metrics: dict[str, dict[str, int]] = {}
    has_market_segments = False

    with attach_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            unit_code = first_value(row, "UNIT_PRODUCTCODE_SF", "unit_code")
            vaps_id = first_value(row, "vaps_item_id", "vaps_code")
            market_segment = first_value(row, "MARKET_SEGMENT_DESCRIPTION")
            has_market_segments = has_market_segments or bool(market_segment)
            if not unit_code or not vaps_id:
                continue

            key = (unit_code, vaps_id)
            rec = recommendations.get(key, {})
            master = vaps_master.get(vaps_id, {})
            attach_rate = (
                number(row.get("attach_rate_2024_26"))
                if "attach_rate_2024_26" in row
                else number(row.get("Vaps_Attach_Rate")) / 100
            )
            vaps_baskets = integer(
                row.get("vaps_baskets_2024_26")
                or row.get("Vaps_Associated_With_Unit")
            )
            unit_baskets = integer(
                row.get("unit_baskets_2024_26") or row.get("Unit_Activations")
            )
            recommended = bool(rec.get("recommended"))
            in_recommendation_sheet = bool(rec.get("in_recommendation_sheet"))
            seen_keys.add(key)
            unit_basket_metrics.setdefault(
                unit_code,
                {
                    "unit_baskets_2024": integer(row.get("unit_baskets_2024")),
                    "unit_baskets_2025": integer(row.get("unit_baskets_2025")),
                    "unit_baskets_2026": integer(row.get("unit_baskets_2026")),
                    "unit_baskets_2024_26": unit_baskets,
                },
            )

            status = status_for_row(
                recommended, attach_rate, vaps_baskets, DEFAULT_ATTACH_RATE_CUTOFF
            )
            action, action_reason = suggested_action(
                recommended, attach_rate, vaps_baskets, DEFAULT_ATTACH_RATE_CUTOFF
            )

            rows.append(
                {
                    "unit_code": unit_code,
                    "market_segment": market_segment,
                    "unit_name": first_value(row, "UNIT_PRODUCTNAME_SF")
                    or rec.get("unit_name", ""),
                    "unit_description": first_value(row, "UNIT_DESCRIPTION")
                    or rec.get("unit_description", ""),
                    "l2": first_value(row, "UNIT_L2_CORE_SOLUTION")
                    or rec.get("unit_l2_core_solution", ""),
                    "l3": first_value(row, "UNIT_L3_PRODUCTS")
                    or rec.get("unit_l3_products", ""),
                    "vaps_id": vaps_id,
                    "vaps_desc": (
                        master.get("description")
                        or first_value(row, "VAPS_DESCRIPTION")
                        or rec.get("vaps_desc")
                        or vaps_lookup.get(vaps_id, "")
                    ),
                    "vaps_source": master.get("source")
                    or first_value(row, "VAPS_SOURCE"),
                    "vaps_main_group": master.get("main_group")
                    or first_value(row, "VAPS_MAIN_GROUP"),
                    "vaps_detailed_group": master.get("detailed_group")
                    or first_value(row, "VAPS_DETAILED_GROUP"),
                    "vaps_package_tier": master.get("package_tier")
                    or first_value(row, "VAPS_PACKAGE_TIER"),
                    "vaps_l1_purpose": master.get("l1_purpose")
                    or first_value(row, "VAPS_L1_PURPOSE"),
                    "vaps_l2_core_need": master.get("l2_core_need")
                    or first_value(row, "VAPS_L2__CORE_NEED"),
                    "recommendation_value": rec.get("recommendation_value", ""),
                    "in_recommendation_sheet": in_recommendation_sheet,
                    "recommended": recommended,
                    "status": status,
                    "suggested_action": action,
                    "action_reason": action_reason,
                    "unit_baskets_2024": integer(row.get("unit_baskets_2024")),
                    "vaps_baskets_2024": integer(row.get("vaps_baskets_2024")),
                    "attach_rate_2024": number(row.get("attach_rate_2024")),
                    "unit_baskets_2025": integer(row.get("unit_baskets_2025")),
                    "vaps_baskets_2025": integer(row.get("vaps_baskets_2025")),
                    "attach_rate_2025": number(row.get("attach_rate_2025")),
                    "unit_baskets_2026": integer(row.get("unit_baskets_2026")),
                    "vaps_baskets_2026": integer(row.get("vaps_baskets_2026")),
                    "attach_rate_2026": number(row.get("attach_rate_2026")),
                    "unit_baskets_2024_26": unit_baskets,
                    "vaps_baskets_2024_26": vaps_baskets,
                    "attach_rate_2024_26": attach_rate,
                    "missed_opportunity_rate": ""
                    if in_recommendation_sheet
                    else attach_rate,
                }
            )

    if has_market_segments:
        return rows

    for (unit_code, vaps_id), rec in sorted(recommendations.items()):
        if (unit_code, vaps_id) in seen_keys:
            continue

        master = vaps_master.get(vaps_id, {})
        recommended = bool(rec.get("recommended"))
        metrics = unit_basket_metrics.get(
            unit_code,
            {
                "unit_baskets_2024": 0,
                "unit_baskets_2025": 0,
                "unit_baskets_2026": 0,
                "unit_baskets_2024_26": 0,
            },
        )
        action, action_reason = suggested_action(
            recommended, 0, 0, DEFAULT_ATTACH_RATE_CUTOFF
        )

        rows.append(
            {
                "unit_code": unit_code,
                "market_segment": "",
                "unit_name": rec.get("unit_name", ""),
                "unit_description": rec.get("unit_description", ""),
                "l2": rec.get("unit_l2_core_solution", ""),
                "l3": rec.get("unit_l3_products", ""),
                "vaps_id": vaps_id,
                "vaps_desc": (
                    master.get("description")
                    or rec.get("vaps_desc")
                    or vaps_lookup.get(vaps_id, "")
                ),
                "vaps_source": master.get("source", ""),
                "vaps_main_group": master.get("main_group", ""),
                "vaps_detailed_group": master.get("detailed_group", ""),
                "vaps_package_tier": master.get("package_tier", ""),
                "vaps_l1_purpose": master.get("l1_purpose", ""),
                "vaps_l2_core_need": master.get("l2_core_need", ""),
                "recommendation_value": rec.get("recommendation_value", ""),
                "in_recommendation_sheet": True,
                "recommended": recommended,
                "status": "Recommended Low Attach" if recommended else "Not Recommended",
                "suggested_action": action,
                "action_reason": action_reason,
                "unit_baskets_2024": metrics["unit_baskets_2024"],
                "vaps_baskets_2024": 0,
                "attach_rate_2024": 0.0,
                "unit_baskets_2025": metrics["unit_baskets_2025"],
                "vaps_baskets_2025": 0,
                "attach_rate_2025": 0.0,
                "unit_baskets_2026": metrics["unit_baskets_2026"],
                "vaps_baskets_2026": 0,
                "attach_rate_2026": 0.0,
                "unit_baskets_2024_26": metrics["unit_baskets_2024_26"],
                "vaps_baskets_2024_26": 0,
                "attach_rate_2024_26": 0.0,
                "missed_opportunity_rate": "",
            }
        )

    return rows


def calculate_elbow_cutoffs(rows: list[dict[str, object]]) -> dict[str, float]:
    rates_by_unit: dict[str, list[float]] = defaultdict(list)

    for row in rows:
        attach_rate = float(row["attach_rate_2024_26"])
        vaps_baskets = int(row["vaps_baskets_2024_26"])
        if attach_rate > 0 and vaps_baskets >= MIN_MISSED_VAPS_BASKETS:
            rates_by_unit[str(row["unit_code"])].append(attach_rate)

    cutoffs: dict[str, float] = {}
    for unit_code, rates in rates_by_unit.items():
        rates = sorted(set(rates), reverse=True)
        if len(rates) < 3:
            cutoffs[unit_code] = rates[-1] if rates else DEFAULT_ATTACH_RATE_CUTOFF
            continue

        first_y = rates[0]
        last_y = rates[-1]
        denominator = ((last_y - first_y) ** 2 + 1) ** 0.5
        best_index = 0
        best_distance = -1.0

        for index, rate in enumerate(rates):
            x = index / (len(rates) - 1)
            distance = abs((last_y - first_y) * x - rate + first_y) / denominator
            if distance > best_distance:
                best_distance = distance
                best_index = index

        cutoffs[unit_code] = rates[best_index]

    return cutoffs


def apply_elbow_decisions(
    rows: list[dict[str, object]],
    cutoffs: dict[str, float],
) -> None:
    for row in rows:
        cutoff = cutoffs.get(str(row["unit_code"]), DEFAULT_ATTACH_RATE_CUTOFF)
        recommended = bool(row["recommended"])
        attach_rate = float(row["attach_rate_2024_26"])
        vaps_baskets = int(row["vaps_baskets_2024_26"])
        action, action_reason = suggested_action(
            recommended, attach_rate, vaps_baskets, cutoff
        )

        row["elbow_cutoff"] = cutoff
        row["status"] = status_for_row(recommended, attach_rate, vaps_baskets, cutoff)
        row["suggested_action"] = action
        row["action_reason"] = action_reason
        row["missed_opportunity_rate"] = (
            "" if row.get("in_recommendation_sheet") else attach_rate
        )


def unit_summary(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        unit_code = str(row["unit_code"])
        summary = grouped.setdefault(
            unit_code,
            {
                "unit_code": unit_code,
                "unit_name": row["unit_name"],
                "unit_baskets": row["unit_baskets_2024_26"],
                "recommended_active": 0,
                "recommended_low": 0,
                "missed": 0,
                "best_missed_rate": 0.0,
            },
        )

        if row["status"] == "Recommended Active":
            summary["recommended_active"] = int(summary["recommended_active"]) + 1
        elif row["status"] == "Recommended Low Attach":
            summary["recommended_low"] = int(summary["recommended_low"]) + 1
        elif row["status"] == "Missed Opportunity":
            summary["missed"] = int(summary["missed"]) + 1
            summary["best_missed_rate"] = max(
                float(summary["best_missed_rate"]), float(row["attach_rate_2024_26"])
            )

    return sorted(
        grouped.values(),
        key=lambda item: (int(item["missed"]), float(item["best_missed_rate"])),
        reverse=True,
    )


def render_rows_for_js(rows: list[dict[str, object]]) -> str:
    compact_rows = []
    for row in rows:
        compact_rows.append(
            {
                "unit": row["unit_code"],
                "market": row.get("market_segment", ""),
                "unitName": row["unit_name"],
                "vaps": row["vaps_id"],
                "vapsDesc": row["vaps_desc"],
                "source": row["vaps_source"],
                "mainGroup": row["vaps_main_group"],
                "detailedGroup": row["vaps_detailed_group"],
                "tier": row["vaps_package_tier"],
                "rec": "Yes" if row["in_recommendation_sheet"] else "No",
                "actionableRec": "Yes" if row["recommended"] else "No",
                "recValue": row["recommendation_value"],
                "status": row["status"],
                "action": row["suggested_action"],
                "reason": row["action_reason"],
                "unitBaskets": row["unit_baskets_2024_26"],
                "vapsBaskets": row["vaps_baskets_2024_26"],
                "attachRate": row["attach_rate_2024_26"],
                "elbowCutoff": row.get("elbow_cutoff", DEFAULT_ATTACH_RATE_CUTOFF),
                "missedRate": row["missed_opportunity_rate"],
                "a2024": row["attach_rate_2024"],
                "a2025": row["attach_rate_2025"],
                "a2026": row["attach_rate_2026"],
            }
        )
    return json.dumps(compact_rows)


def write_report(rows: list[dict[str, object]], output_path: Path) -> None:
    table_rows = sorted(
        rows,
        key=lambda item: (
            item["status"] != "Missed Opportunity",
            -float(item["attach_rate_2024_26"]),
            -int(item["vaps_baskets_2024_26"]),
        ),
    )

    js_rows = render_rows_for_js(table_rows)
    units = sorted({str(row["unit_code"]) for row in rows})
    market_segments = sorted(
        {str(row.get("market_segment") or "Unmapped") for row in rows if row.get("market_segment")}
    )
    has_market_segments = bool(market_segments)
    sources = sorted({str(row["vaps_source"] or "Unmapped") for row in rows})
    main_groups = sorted({str(row["vaps_main_group"] or "Unmapped") for row in rows})
    tiers = sorted({str(row["vaps_package_tier"] or "Unmapped") for row in rows})

    unit_options = "\n".join(
        f'<option value="{html.escape(unit)}"{" selected" if unit == "10W" else ""}>{html.escape(unit)}</option>'
        for unit in units
    )
    source_options = "\n".join(
        f'<option value="{html.escape(source)}">{html.escape(source)}</option>'
        for source in sources
    )
    market_options = "\n".join(
        f'<option value="{html.escape(market)}">{html.escape(market)}</option>'
        for market in market_segments
    )
    group_options = "\n".join(
        f'<option value="{html.escape(group)}">{html.escape(group)}</option>'
        for group in main_groups
    )
    tier_options = "\n".join(
        f'<option value="{html.escape(tier)}">{html.escape(tier)}</option>' for tier in tiers
    )

    market_filter_html = (
        f"""        <label>Market Segment
          <select id="marketFilter">
            <option value="">All market segments</option>
            {market_options}
          </select>
        </label>
"""
        if has_market_segments
        else ""
    )
    market_header_html = (
        '<th class="sortable" data-sort="market">Market segment</th>'
        if has_market_segments
        else ""
    )
    market_column_js = (
        '${row.market ? `<td>${esc(row.market || "Unmapped")}</td>` : ""}'
        if has_market_segments
        else ""
    )
    market_filter_js = (
        '    const marketFilter = document.getElementById("marketFilter");'
        if has_market_segments
        else '    const marketFilter = null;'
    )
    market_filter_value_js = (
        '      const market = marketFilter ? marketFilter.value : "";'
        if has_market_segments
        else '      const market = "";'
    )
    market_filter_condition_js = (
        '        if (market && (row.market || "Unmapped") !== market) return false;'
        if has_market_segments
        else ""
    )
    market_export_column_js = (
        '        ["Market Segment", "market"],\n'
        if has_market_segments
        else ""
    )
    market_controls_js = (
        "marketFilter, "
        if has_market_segments
        else ""
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>VAPS Attach Rate Report</title>
  <style>
    :root {{
      --ink: #17212b;
      --muted: #62707f;
      --line: #d9e0e8;
      --paper: #f7f9fb;
      --panel: #ffffff;
      --green: #168a5b;
      --red: #b23b3b;
      --gold: #a56b00;
      --blue: #2563a7;
      --teal: #147a86;
    }}
    * {{ box-sizing: border-box; }}
    html {{
      height: 100%;
      overflow: hidden;
    }}
    body {{
      margin: 0;
      height: 100%;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: var(--paper);
      letter-spacing: 0;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }}
    main {{
      padding: 24px 32px;
      display: grid;
      gap: 24px;
      flex: 1 1 auto;
      min-height: 0;
      overflow: hidden;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 20px;
      display: flex;
      flex-direction: column;
      min-width: 0;
      min-height: 0;
      overflow: hidden;
    }}
    h2 {{
      margin: 0 0 14px;
      font-size: 20px;
    }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(6, minmax(150px, 1fr));
      gap: 12px;
    }}
    .kpi {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      background: #fbfcfd;
      min-height: 98px;
    }}
    .kpi span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
      line-height: 1.3;
      margin-bottom: 10px;
    }}
    .kpi strong {{
      display: block;
      font-size: 26px;
      line-height: 1.1;
    }}
    .kpi small {{
      display: block;
      margin-top: 6px;
      color: var(--muted);
      line-height: 1.35;
    }}
    .grid-two {{
      display: grid;
      grid-template-columns: minmax(340px, 1fr) minmax(340px, 1fr);
      gap: 24px;
    }}
    .bar-list {{
      display: grid;
      gap: 10px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(220px, 1.6fr) minmax(140px, 2fr) 70px;
      gap: 12px;
      align-items: center;
    }}
    .bar-label strong,
    .bar-label span {{
      display: block;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .bar-label span {{
      color: var(--muted);
      font-size: 12px;
      margin-top: 2px;
    }}
    .bar-track {{
      height: 14px;
      background: #edf1f5;
      border-radius: 8px;
      overflow: hidden;
      border: 1px solid #d5dde5;
    }}
    .bar-fill {{
      height: 100%;
      background: linear-gradient(90deg, var(--teal), var(--green));
    }}
    .bar-value {{
      text-align: right;
      font-weight: 700;
    }}
    .controls {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 12px;
      margin-bottom: 14px;
      flex: 0 0 auto;
      align-items: end;
      min-width: 0;
    }}
    .controls label {{
      min-width: 0;
    }}
    .controls label:last-child {{
      min-width: min(100%, 260px);
    }}
    .toolbar {{
      display: flex;
      justify-content: flex-end;
      margin-bottom: 10px;
      flex: 0 0 auto;
      min-width: 0;
    }}
    button {{
      border: 1px solid #b8c4d0;
      border-radius: 6px;
      background: #ffffff;
      color: var(--ink);
      cursor: pointer;
      font-size: 14px;
      font-weight: 700;
      padding: 9px 12px;
    }}
    button:hover {{
      background: #eef3f8;
    }}
    .content-split {{
      display: grid;
      grid-template-columns: minmax(170px, var(--chart-width, 240px)) 8px minmax(0, 1fr);
      gap: 10px;
      flex: 1 1 auto;
      min-height: 0;
      min-width: 0;
    }}
    .splitter {{
      cursor: col-resize;
      border-radius: 6px;
      background: #d9e0e8;
      min-height: 100%;
      position: relative;
    }}
    .splitter::after {{
      content: "";
      position: absolute;
      top: 50%;
      left: 50%;
      width: 2px;
      height: 44px;
      transform: translate(-50%, -50%);
      border-left: 1px solid #8d99a6;
      border-right: 1px solid #8d99a6;
    }}
    .splitter:hover,
    .splitter.active {{
      background: #b8c4d0;
    }}
    label {{
      display: grid;
      gap: 5px;
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
    }}
    input,
    select {{
      width: 100%;
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 9px 10px;
      font-size: 14px;
      color: var(--ink);
      background: #fff;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      height: 100%;
      min-height: 0;
      max-width: 100%;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 1620px;
      background: #fff;
    }}
    th,
    td {{
      border-bottom: 1px solid var(--line);
      padding: 10px 12px;
      text-align: left;
      vertical-align: top;
      font-size: 13px;
      line-height: 1.35;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #eef3f8;
      z-index: 1;
      font-size: 12px;
      text-transform: uppercase;
    }}
    th.sortable {{
      cursor: pointer;
      user-select: none;
      white-space: nowrap;
    }}
    th.sortable::after {{
      content: " ↕";
      color: var(--muted);
      font-weight: 400;
    }}
    th.sortable.sort-asc::after {{
      content: " ↑";
      color: var(--ink);
      font-weight: 700;
    }}
    th.sortable.sort-desc::after {{
      content: " ↓";
      color: var(--ink);
      font-weight: 700;
    }}
    .num {{ text-align: right; white-space: nowrap; }}
    .emph {{ color: var(--red); font-weight: 700; }}
    .pill {{
      display: inline-block;
      border-radius: 6px;
      padding: 3px 7px;
      font-weight: 700;
      font-size: 12px;
      white-space: nowrap;
    }}
    .missed {{ background: #fdecec; color: var(--red); }}
    .active {{ background: #e8f6ef; color: var(--green); }}
    .low {{ background: #fff3d9; color: var(--gold); }}
    .none {{ background: #eef1f4; color: #53606d; }}
    .keep {{ background: #e8f6ef; color: var(--green); }}
    .add {{ background: #e8f1fb; color: var(--blue); }}
    .remove {{ background: #fdecec; color: var(--red); }}
    .review {{ background: #fff3d9; color: var(--gold); }}
    .monitor {{ background: #eef1f4; color: #53606d; }}
    .note {{
      color: var(--muted);
      line-height: 1.5;
      margin: 10px 0 0;
      flex: 0 0 auto;
    }}
    .chart-panel {{
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-bottom: 0;
      padding: 12px;
      min-width: 0;
      min-height: 0;
      background: #fbfcfd;
      display: flex;
      flex-direction: column;
      overflow: hidden;
      position: relative;
    }}
    .chart-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: baseline;
      margin-bottom: 8px;
    }}
    .chart-head strong {{
      font-size: 14px;
    }}
    .chart-head span {{
      color: var(--muted);
      font-size: 12px;
    }}
    #elbowChart {{
      width: 100%;
      height: 100%;
      display: block;
      min-height: 0;
    }}
    .chart-tooltip {{
      position: absolute;
      display: none;
      pointer-events: none;
      z-index: 3;
      max-width: 240px;
      border: 1px solid #cdd7e1;
      border-radius: 8px;
      background: #ffffff;
      box-shadow: 0 8px 22px rgba(23, 33, 43, 0.14);
      padding: 9px 10px;
      color: var(--ink);
      font-size: 12px;
      line-height: 1.35;
    }}
    .chart-tooltip strong {{
      display: block;
      margin-bottom: 4px;
    }}
    .chart-tooltip span {{
      display: block;
      color: var(--muted);
    }}
    @media (max-width: 1100px) {{
      .kpis {{ grid-template-columns: repeat(3, minmax(150px, 1fr)); }}
      .grid-two {{ grid-template-columns: 1fr; }}
      .controls {{ grid-template-columns: repeat(auto-fit, minmax(170px, 1fr)); }}
      .content-split {{
        grid-template-columns: 1fr;
        grid-template-rows: 180px minmax(0, 1fr);
      }}
      .splitter {{ display: none; }}
    }}
    @media (max-width: 900px) {{
      html {{
        height: auto;
        min-height: 100%;
        overflow: auto;
      }}
      body {{
        height: auto;
        min-height: 100%;
        overflow: auto;
      }}
      main {{
        padding: 18px 20px 24px;
        overflow: visible;
      }}
      section {{
        min-height: 760px;
        overflow: visible;
      }}
      .content-split {{
        min-height: 560px;
      }}
      .table-wrap {{
        max-height: 520px;
      }}
      .toolbar {{
        justify-content: stretch;
      }}
      button {{
        width: 100%;
      }}
    }}
    @media (max-width: 700px) {{
      main {{ padding-left: 16px; padding-right: 16px; }}
      .kpis, .controls {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 1fr; gap: 6px; }}
      .bar-value {{ text-align: left; }}
      section {{
        padding: 14px;
        min-height: 720px;
      }}
      .content-split {{
        grid-template-rows: 170px minmax(0, 1fr);
        min-height: 520px;
      }}
      .chart-panel {{ min-height: 0; }}
      .chart-head {{
        display: grid;
        gap: 4px;
      }}
      table {{
        min-width: 1500px;
      }}
    }}
    @media (max-width: 480px) {{
      main {{
        padding-top: 14px;
        padding-bottom: 18px;
      }}
      h2 {{
        font-size: 18px;
      }}
      input,
      select,
      button {{
        font-size: 16px;
      }}
    }}
  </style>
</head>
<body>
  <main>
    <section>
      <h2>Recommendation Decision Table</h2>
      <div class="controls">
        <label>Unit
          <select id="unitFilter">
            <option value="">All units</option>
            {unit_options}
          </select>
        </label>
{market_filter_html.rstrip()}
        <label>Status
          <select id="statusFilter">
            <option value="">All statuses</option>
            <option>Missed Opportunity</option>
            <option>Recommended Meets Cutoff</option>
            <option>Recommended Below Cutoff</option>
            <option>Below Cutoff</option>
          </select>
        </label>
        <label>Action
          <select id="actionFilter">
            <option value="">All actions</option>
            <option>Add</option>
            <option>Keep</option>
            <option>Monitor</option>
            <option>No Action</option>
            <option>Remove</option>
          </select>
        </label>
        <label>Current Rec
          <select id="recommendationFilter">
            <option value="">All</option>
            <option value="Yes">Yes</option>
            <option value="No">No</option>
          </select>
        </label>
        <label>VAPS Source
          <select id="sourceFilter">
            <option value="">All sources</option>
            {source_options}
          </select>
        </label>
        <label>Main Group
          <select id="groupFilter">
            <option value="">All groups</option>
            {group_options}
          </select>
        </label>
        <label>Package Tier
          <select id="tierFilter">
            <option value="">All tiers</option>
            {tier_options}
          </select>
        </label>
        <label>Minimum Attach %
          <input id="minAttach" type="number" min="0" max="100" step="0.1" value="0">
        </label>
        <label>Search
          <input id="searchBox" type="search" placeholder="Unit, VAPS, status, recommendation value">
        </label>
      </div>
      <div class="toolbar">
        <button id="downloadCsv" type="button">Download CSV</button>
      </div>
      <div class="content-split">
        <div class="chart-panel">
          <div class="chart-head">
            <strong>Elbow Chart</strong>
            <span id="elbowSummary"></span>
          </div>
          <canvas id="elbowChart" width="1200" height="260"></canvas>
          <div class="chart-tooltip" id="chartTooltip"></div>
        </div>
        <div class="splitter" id="splitter" title="Drag to resize chart and table"></div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th class="sortable" data-sort="unit">Unit</th>
                {market_header_html}
                <th class="sortable" data-sort="unitName">Unit name</th>
                <th class="sortable" data-sort="vaps">VAPS</th>
                <th class="sortable" data-sort="vapsDesc">VAPS description</th>
                <th class="sortable" data-sort="source">Source</th>
                <th class="sortable" data-sort="mainGroup">Main group</th>
                <th class="sortable" data-sort="detailedGroup">Detailed group</th>
                <th class="sortable" data-sort="tier">Tier</th>
                <th class="sortable" data-sort="rec">Current recommendation</th>
                <th class="sortable" data-sort="status">Status</th>
                <th class="sortable" data-sort="action">Suggested action</th>
                <th class="sortable" data-sort="reason">Reason</th>
                <th class="num sortable" data-sort="unitBaskets">Unit baskets</th>
                <th class="num sortable" data-sort="vapsBaskets">VAPS baskets</th>
                <th class="num sortable" data-sort="attachRate">Attach rate</th>
                <th class="num sortable" data-sort="elbowCutoff" title="For each unit, VAPS attach rates with at least 25 baskets are sorted from highest to lowest. The elbow cutoff is the point farthest from the straight line between the highest and lowest attach rates.">Elbow cutoff</th>
                <th class="num sortable" data-sort="missedRate">Missed opportunity rate</th>
                <th class="num sortable" data-sort="a2024">2024</th>
                <th class="num sortable" data-sort="a2025">2025</th>
                <th class="num sortable" data-sort="a2026">2026</th>
              </tr>
            </thead>
            <tbody id="detailBody"></tbody>
          </table>
        </div>
      </div>
      <p class="note" id="rowCount"></p>
    </section>
  </main>
  <script>
    const rows = {js_rows};
    const body = document.getElementById("detailBody");
    const rowCount = document.getElementById("rowCount");
    const unitFilter = document.getElementById("unitFilter");
{market_filter_js}
    const statusFilter = document.getElementById("statusFilter");
    const actionFilter = document.getElementById("actionFilter");
    const recommendationFilter = document.getElementById("recommendationFilter");
    const sourceFilter = document.getElementById("sourceFilter");
    const groupFilter = document.getElementById("groupFilter");
    const tierFilter = document.getElementById("tierFilter");
    const minAttach = document.getElementById("minAttach");
    const searchBox = document.getElementById("searchBox");
    const downloadCsv = document.getElementById("downloadCsv");
    const elbowChart = document.getElementById("elbowChart");
    const elbowSummary = document.getElementById("elbowSummary");
    const chartTooltip = document.getElementById("chartTooltip");
    const contentSplit = document.querySelector(".content-split");
    const splitter = document.getElementById("splitter");
    const sortableHeaders = document.querySelectorAll("th.sortable");
    let chartPoints = [];
    let sortState = {{ key: "rec", direction: "desc" }};

    const savedChartWidth = localStorage.getItem("vapsReportChartWidth");
    if (savedChartWidth) {{
      contentSplit.style.setProperty("--chart-width", savedChartWidth);
    }}

    function esc(value) {{
      return String(value ?? "").replace(/[&<>"']/g, char => ({{
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        '"': "&quot;",
        "'": "&#39;"
      }}[char]));
    }}

    function fmtPct(value) {{
      if (value === "" || value === null || value === undefined) return "";
      return `${{(Number(value || 0) * 100).toFixed(1)}}%`;
    }}

    function fmtInt(value) {{
      return Number(value || 0).toLocaleString();
    }}

    function resizeCanvas(canvas) {{
      const ratio = window.devicePixelRatio || 1;
      const rect = canvas.getBoundingClientRect();
      canvas.width = Math.max(1, Math.floor(rect.width * ratio));
      canvas.height = Math.max(1, Math.floor(rect.height * ratio));
      const ctx = canvas.getContext("2d");
      ctx.setTransform(ratio, 0, 0, ratio, 0, 0);
      return {{ ctx, width: rect.width, height: rect.height }};
    }}

    function drawElbowChart(unit) {{
      const {{ ctx, width, height }} = resizeCanvas(elbowChart);
      chartPoints = [];
      chartTooltip.style.display = "none";
      ctx.clearRect(0, 0, width, height);

      const unitRows = rows
        .filter(row => (!unit || row.unit === unit) && row.attachRate > 0 && row.vapsBaskets >= {MIN_MISSED_VAPS_BASKETS})
        .sort((a, b) => b.attachRate - a.attachRate);

      if (!unitRows.length) {{
        elbowSummary.textContent = "No rows meet the minimum basket count for this selection.";
        ctx.fillStyle = "#62707f";
        ctx.font = "13px Arial";
        ctx.fillText("No elbow chart data available", 16, 32);
        return;
      }}

      const cutoff = unitRows[0].elbowCutoff || 0;
      const cutoffIndex = Math.max(0, unitRows.findIndex(row => Math.abs(row.attachRate - cutoff) < 0.0000005));
      const maxRate = Math.max(...unitRows.map(row => row.attachRate), cutoff, 0.01);
      const left = 46;
      const right = 18;
      const top = 16;
      const bottom = 34;
      const chartWidth = width - left - right;
      const chartHeight = height - top - bottom;
      const xFor = index => left + (unitRows.length === 1 ? 0 : index / (unitRows.length - 1) * chartWidth);
      const yFor = rate => top + chartHeight - (rate / maxRate * chartHeight);

      ctx.strokeStyle = "#d9e0e8";
      ctx.lineWidth = 1;
      ctx.beginPath();
      ctx.moveTo(left, top);
      ctx.lineTo(left, top + chartHeight);
      ctx.lineTo(left + chartWidth, top + chartHeight);
      ctx.stroke();

      ctx.fillStyle = "#62707f";
      ctx.font = "12px Arial";
      ctx.textAlign = "right";
      ctx.fillText(fmtPct(maxRate), left - 8, top + 4);
      ctx.fillText("0.0%", left - 8, top + chartHeight + 4);
      ctx.textAlign = "center";
      ctx.fillText("VAPS ranked by attach rate", left + chartWidth / 2, height - 8);

      ctx.strokeStyle = "#2563a7";
      ctx.lineWidth = 2;
      ctx.beginPath();
      unitRows.forEach((row, index) => {{
        const x = xFor(index);
        const y = yFor(row.attachRate);
        chartPoints.push({{
          x,
          y,
          row,
          index,
          isElbow: index === cutoffIndex
        }});
        if (index === 0) ctx.moveTo(x, y);
        else ctx.lineTo(x, y);
      }});
      ctx.stroke();

      const cutoffY = yFor(cutoff);
      ctx.strokeStyle = "#b23b3b";
      ctx.setLineDash([6, 5]);
      ctx.beginPath();
      ctx.moveTo(left, cutoffY);
      ctx.lineTo(left + chartWidth, cutoffY);
      ctx.stroke();
      ctx.setLineDash([]);

      if (cutoffIndex >= 0) {{
        const x = xFor(cutoffIndex);
        const y = yFor(unitRows[cutoffIndex].attachRate);
        ctx.fillStyle = "#b23b3b";
        ctx.beginPath();
        ctx.arc(x, y, 5, 0, Math.PI * 2);
        ctx.fill();
      }}

      ctx.fillStyle = "#b23b3b";
      ctx.font = "12px Arial";
      ctx.textAlign = "left";
      ctx.fillText(`Elbow cutoff: ${{fmtPct(cutoff)}}`, left + 8, Math.max(top + 12, cutoffY - 8));

      const unitText = unit || "All units";
      elbowSummary.textContent = `${{unitText}} | ${{unitRows.length.toLocaleString()}} VAPS with {MIN_MISSED_VAPS_BASKETS:,}+ baskets | cutoff ${{fmtPct(cutoff)}}`;
    }}

    function showChartTooltip(event) {{
      if (!chartPoints.length) return;
      const rect = elbowChart.getBoundingClientRect();
      const x = event.clientX - rect.left;
      const y = event.clientY - rect.top;
      let best = null;
      let bestDistance = Infinity;

      chartPoints.forEach(point => {{
        const distance = Math.hypot(point.x - x, point.y - y);
        if (distance < bestDistance) {{
          best = point;
          bestDistance = distance;
        }}
      }});

      if (!best || bestDistance > 18) {{
        chartTooltip.style.display = "none";
        return;
      }}

      const row = best.row;
      chartTooltip.innerHTML = `
        <strong>${{best.isElbow ? "Elbow cutoff point" : "Attach rate point"}}</strong>
        <span>Unit: ${{esc(row.unit)}}</span>
        <span>Rank: ${{best.index + 1}} of ${{chartPoints.length}}</span>
        <span>VAPS: ${{esc(row.vaps)}}</span>
        <span>Attach rate: ${{fmtPct(row.attachRate)}}</span>
        <span>VAPS baskets: ${{fmtInt(row.vapsBaskets)}}</span>
        <span>Elbow cutoff: ${{fmtPct(row.elbowCutoff)}}</span>
      `;

      const panelRect = chartTooltip.parentElement.getBoundingClientRect();
      const tooltipWidth = chartTooltip.offsetWidth || 220;
      const tooltipHeight = chartTooltip.offsetHeight || 120;
      let left = event.clientX - panelRect.left + 12;
      let top = event.clientY - panelRect.top + 12;

      if (left + tooltipWidth > panelRect.width - 8) left = event.clientX - panelRect.left - tooltipWidth - 12;
      if (top + tooltipHeight > panelRect.height - 8) top = event.clientY - panelRect.top - tooltipHeight - 12;

      chartTooltip.style.left = `${{Math.max(8, left)}}px`;
      chartTooltip.style.top = `${{Math.max(8, top)}}px`;
      chartTooltip.style.display = "block";
    }}

    elbowChart.addEventListener("mousemove", showChartTooltip);
    elbowChart.addEventListener("mouseleave", () => {{
      chartTooltip.style.display = "none";
    }});

    function pillClass(status) {{
      if (status === "Missed Opportunity") return "pill missed";
      if (status === "Recommended Meets Cutoff") return "pill active";
      if (status === "Recommended Below Cutoff") return "pill low";
      if (status === "Below Cutoff") return "pill none";
      return "pill none";
    }}

    function actionClass(action) {{
      if (action === "Keep") return "pill keep";
      if (action === "Add") return "pill add";
      if (action === "Remove") return "pill remove";
      if (action === "Monitor") return "pill monitor";
      return "pill none";
    }}

    function sortValue(row, key) {{
      const value = row[key];
      if (value === "" || value === null || value === undefined) return null;
      if (typeof value === "number") return value;
      return String(value).toLowerCase();
    }}

    function applySort(inputRows) {{
      const {{ key, direction }} = sortState;
      const factor = direction === "asc" ? 1 : -1;

      return [...inputRows].sort((a, b) => {{
        const av = sortValue(a, key);
        const bv = sortValue(b, key);
        if (av === null && bv === null) return secondarySort(a, b);
        if (av === null) return 1;
        if (bv === null) return -1;
        let result = 0;
        if (typeof av === "number" && typeof bv === "number") {{
          result = (av - bv) * factor;
        }} else {{
          result = String(av).localeCompare(String(bv), undefined, {{ numeric: true }}) * factor;
        }}
        return result || secondarySort(a, b);
      }});
    }}

    function secondarySort(a, b) {{
      return (
        (b.attachRate - a.attachRate) ||
        (b.vapsBaskets - a.vapsBaskets) ||
        String(a.vaps).localeCompare(String(b.vaps), undefined, {{ numeric: true }})
      );
    }}

    function updateSortIndicators() {{
      sortableHeaders.forEach(header => {{
        header.classList.remove("sort-asc", "sort-desc");
        if (header.dataset.sort === sortState.key) {{
          header.classList.add(sortState.direction === "asc" ? "sort-asc" : "sort-desc");
        }}
      }});
    }}

    function elbowTooltip(row) {{
      return `Elbow cutoff for ${{row.unit}}: ${{fmtPct(row.elbowCutoff)}}. Calculated by sorting this unit's VAPS attach rates from highest to lowest, keeping VAPS with at least {MIN_MISSED_VAPS_BASKETS} baskets, drawing a straight line from the highest to lowest attach rate, and selecting the point farthest from that line.`;
    }}

    function filteredRows() {{
      const unit = unitFilter.value;
{market_filter_value_js}
      const status = statusFilter.value;
      const action = actionFilter.value;
      const recommendation = recommendationFilter.value;
      const source = sourceFilter.value;
      const group = groupFilter.value;
      const tier = tierFilter.value;
      const minRate = Number(minAttach.value || 0) / 100;
      const query = searchBox.value.trim().toLowerCase();

      return applySort(rows.filter(row => {{
        if (unit && row.unit !== unit) return false;
{market_filter_condition_js}
        if (status && row.status !== status) return false;
        if (action && row.action !== action) return false;
        if (recommendation && row.rec !== recommendation) return false;
        if (source && (row.source || "Unmapped") !== source) return false;
        if (group && (row.mainGroup || "Unmapped") !== group) return false;
        if (tier && (row.tier || "Unmapped") !== tier) return false;
        if (row.attachRate < minRate) return false;
        if (!query) return true;
        return [
          row.unit,
          row.market,
          row.unitName,
          row.vaps,
          row.vapsDesc,
          row.source,
          row.mainGroup,
          row.detailedGroup,
          row.tier,
          row.rec,
          row.recValue,
          row.status,
          row.action,
          row.reason
        ].join(" ").toLowerCase().includes(query);
      }}));
    }}

    function render() {{
      const unit = unitFilter.value;
      const filtered = filteredRows();

      body.innerHTML = filtered.slice(0, 500).map(row => `
        <tr>
          <td>${{esc(row.unit)}}</td>
          {market_column_js}
          <td>${{esc(row.unitName)}}</td>
          <td><strong>${{esc(row.vaps)}}</strong></td>
          <td>${{esc(row.vapsDesc || "No description in recommendation sheet")}}</td>
          <td>${{esc(row.source || "Unmapped")}}</td>
          <td>${{esc(row.mainGroup || "Unmapped")}}</td>
          <td>${{esc(row.detailedGroup || "Unmapped")}}</td>
          <td>${{esc(row.tier || "Unmapped")}}</td>
          <td>${{esc(row.rec)}}${{row.recValue ? `<br><small>${{esc(row.recValue)}}</small>` : ""}}</td>
          <td><span class="${{pillClass(row.status)}}">${{esc(row.status)}}</span></td>
          <td><span class="${{actionClass(row.action)}}">${{esc(row.action)}}</span></td>
          <td>${{esc(row.reason)}}</td>
          <td class="num">${{fmtInt(row.unitBaskets)}}</td>
          <td class="num">${{fmtInt(row.vapsBaskets)}}</td>
          <td class="num"><strong>${{fmtPct(row.attachRate)}}</strong></td>
          <td class="num" title="${{esc(elbowTooltip(row))}}">${{fmtPct(row.elbowCutoff)}}</td>
          <td class="num">${{fmtPct(row.missedRate)}}</td>
          <td class="num">${{fmtPct(row.a2024)}}</td>
          <td class="num">${{fmtPct(row.a2025)}}</td>
          <td class="num">${{fmtPct(row.a2026)}}</td>
        </tr>
      `).join("");

      rowCount.textContent = `Showing ${{Math.min(filtered.length, 500).toLocaleString()}} of ${{filtered.length.toLocaleString()}} matching rows.`;
      drawElbowChart(unit);
    }}

    function csvEscape(value) {{
      const text = String(value ?? "");
      if (/[",\\r\\n]/.test(text)) {{
        return `"${{text.replace(/"/g, '""')}}"`;
      }}
      return text;
    }}

    function downloadFilteredCsv() {{
      const exportRows = filteredRows();
      const columns = [
        ["Unit", "unit"],
{market_export_column_js.rstrip()}
        ["Unit Name", "unitName"],
        ["VAPS", "vaps"],
        ["VAPS Description", "vapsDesc"],
        ["Source", "source"],
        ["Main Group", "mainGroup"],
        ["Detailed Group", "detailedGroup"],
        ["Tier", "tier"],
        ["Current Recommendation", "rec"],
        ["Recommendation Value", "recValue"],
        ["Status", "status"],
        ["Suggested Action", "action"],
        ["Reason", "reason"],
        ["Unit Baskets", "unitBaskets"],
        ["VAPS Baskets", "vapsBaskets"],
        ["Attach Rate", "attachRate"],
        ["Elbow Cutoff", "elbowCutoff"],
        ["Missed Opportunity Rate", "missedRate"],
        ["Attach Rate 2024", "a2024"],
        ["Attach Rate 2025", "a2025"],
        ["Attach Rate 2026", "a2026"]
      ];

      const percentFields = new Set(["attachRate", "elbowCutoff", "missedRate", "a2024", "a2025", "a2026"]);
      const csvRows = [
        columns.map(([label]) => csvEscape(label)).join(","),
        ...exportRows.map(row => columns.map(([, key]) => {{
          const value = row[key];
          if (percentFields.has(key)) return csvEscape(value === "" ? "" : fmtPct(value));
          return csvEscape(value);
        }}).join(","))
      ];

      const blob = new Blob(["\ufeff" + csvRows.join("\\r\\n")], {{ type: "text/csv;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const unit = unitFilter.value || "all_units";
      const stamp = new Date().toISOString().slice(0, 10);
      const link = document.createElement("a");
      link.href = url;
      link.download = `vaps_recommendation_decision_${{unit}}_${{stamp}}.csv`;
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    }}

    [unitFilter, {market_controls_js}statusFilter, actionFilter, recommendationFilter, sourceFilter, groupFilter, tierFilter, minAttach, searchBox].forEach(control => {{
      if (!control) return;
      control.addEventListener("input", render);
      control.addEventListener("change", render);
    }});
    downloadCsv.addEventListener("click", downloadFilteredCsv);
    sortableHeaders.forEach(header => {{
      header.addEventListener("click", () => {{
        const key = header.dataset.sort;
        if (sortState.key === key) {{
          sortState.direction = sortState.direction === "asc" ? "desc" : "asc";
        }} else {{
          sortState.key = key;
          sortState.direction = header.classList.contains("num") ? "desc" : "asc";
        }}
        updateSortIndicators();
        render();
      }});
    }});

    let resizing = false;

    splitter.addEventListener("pointerdown", event => {{
      resizing = true;
      splitter.classList.add("active");
      splitter.setPointerCapture(event.pointerId);
      document.body.style.userSelect = "none";
    }});

    splitter.addEventListener("pointermove", event => {{
      if (!resizing) return;
      const rect = contentSplit.getBoundingClientRect();
      const minWidth = 170;
      const maxWidth = Math.max(minWidth, rect.width * 0.55);
      const nextWidth = Math.min(maxWidth, Math.max(minWidth, event.clientX - rect.left));
      const value = `${{Math.round(nextWidth)}}px`;
      contentSplit.style.setProperty("--chart-width", value);
      localStorage.setItem("vapsReportChartWidth", value);
      drawElbowChart(unitFilter.value);
    }});

    function stopResize(event) {{
      if (!resizing) return;
      resizing = false;
      splitter.classList.remove("active");
      document.body.style.userSelect = "";
      if (event.pointerId !== undefined) {{
        try {{ splitter.releasePointerCapture(event.pointerId); }} catch {{}}
      }}
      drawElbowChart(unitFilter.value);
    }}

    splitter.addEventListener("pointerup", stopResize);
    splitter.addEventListener("pointercancel", stopResize);

    updateSortIndicators();
    render();
    window.addEventListener("resize", () => drawElbowChart(unitFilter.value));
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )


def generate_report(attach_csv: Path, output_html: Path) -> None:
    recommendations, vaps_lookup = load_recommendations(RECOMMENDATION_XLSX)
    vaps_master = load_vaps_master(VAPS_MASTER_CSV)
    rows = load_attach_rows(attach_csv, recommendations, vaps_lookup, vaps_master)
    cutoffs = calculate_elbow_cutoffs(rows)
    apply_elbow_decisions(rows, cutoffs)
    write_report(rows, output_html)
    print(f"Wrote {output_html}")
    print(f"Source CSV: {attach_csv}")
    print(f"Rows in report: {len(rows):,}")
    print(f"VAPS master rows: {len(vaps_master):,}")
    print(f"Unit elbow cutoffs calculated: {len(cutoffs):,}")


def main() -> None:
    generate_report(ATTACH_RATE_CSV, REPORT_HTML)
    generate_report(MARKET_ATTACH_RATE_CSV, MARKET_REPORT_HTML)


if __name__ == "__main__":
    main()
