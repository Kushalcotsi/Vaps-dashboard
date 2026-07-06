import os
from typing import List, Dict, Tuple, Optional
import snowflake.connector
from snowflake.connector.cursor import SnowflakeCursor
import re
import threading
from functools import lru_cache
from app.repositories.base import BaseRepository
from app.models.dashboard import VapsAttachRate, RecommendationEntry
from app.core.config import settings
from openpyxl import load_workbook

class SnowflakeRepository(BaseRepository):
    def __init__(self):
        self.recommendations = self._load_recommendations()
        self._conn = None
        self._lock = threading.Lock()

    def _get_connection(self):
        # If we have a connection, try to reuse it
        if self._conn is not None:
            try:
                if not self._conn.is_closed():
                    return self._conn
            except Exception:
                self._conn = None

        with self._lock:
            # Re-check inside lock
            if self._conn is not None:
                try:
                    if not self._conn.is_closed():
                        return self._conn
                except:
                    pass

            print(f"Connecting to Snowflake as {settings.SNOWFLAKE_USER} (Auth: {settings.SNOWFLAKE_AUTHENTICATOR})...")
            
            # Setup connection arguments with timeouts
            conn_args = {
                "user": settings.SNOWFLAKE_USER,
                "account": settings.SNOWFLAKE_ACCOUNT,
                "warehouse": settings.SNOWFLAKE_WAREHOUSE,
                "database": settings.SNOWFLAKE_DATABASE,
                "schema": settings.SNOWFLAKE_SCHEMA,
                "role": settings.SNOWFLAKE_ROLE,
                "login_timeout": settings.SNOWFLAKE_CONNECTION_TIMEOUT,
                "network_timeout": settings.SNOWFLAKE_CONNECTION_TIMEOUT,
                "socket_timeout": settings.SNOWFLAKE_CONNECTION_TIMEOUT
            }
            
            
            authenticator = (settings.SNOWFLAKE_AUTHENTICATOR or "snowflake").lower()
            
            if authenticator == "externalbrowser":
                print("⚠️ WARNING: 'externalbrowser' SSO authenticator is active. In a headless environment (e.g. EC2 under PM2), this will block the process waiting for interactive stdin redirect URLs.")
                conn_args["authenticator"] = "externalbrowser"
                
            elif authenticator == "keypair" or settings.SNOWFLAKE_PRIVATE_KEY_PATH:
                print(f"Loading private key from path: {settings.SNOWFLAKE_PRIVATE_KEY_PATH}...")
                try:
                    from cryptography.hazmat.primitives import serialization
                    
                    with open(settings.SNOWFLAKE_PRIVATE_KEY_PATH, "rb") as key_file:
                        passphrase = settings.SNOWFLAKE_PRIVATE_KEY_PASSPHRASE.encode() if settings.SNOWFLAKE_PRIVATE_KEY_PASSPHRASE else None
                        private_key = serialization.load_pem_private_key(
                            key_file.read(),
                            password=passphrase
                        )
                        
                    private_key_bytes = private_key.private_bytes(
                        encoding=serialization.Encoding.DER,
                        format=serialization.PrivateFormat.PKCS8,
                        encryption_algorithm=serialization.NoEncryption()
                    )
                    conn_args["private_key"] = private_key_bytes
                except Exception as e:
                    print(f"❌ Error loading or parsing private key: {str(e)}")
                    raise e
                    
            else:
                # Default username/password authentication
                if not settings.SNOWFLAKE_PASSWORD:
                    print("⚠️ WARNING: Snowflake password is empty but using standard password authentication.")
                conn_args["password"] = settings.SNOWFLAKE_PASSWORD
                # If authenticator is explicitly set and not empty, pass it.
                if settings.SNOWFLAKE_AUTHENTICATOR and settings.SNOWFLAKE_AUTHENTICATOR.lower() != "snowflake":
                    conn_args["authenticator"] = settings.SNOWFLAKE_AUTHENTICATOR

            try:
                self._conn = snowflake.connector.connect(**conn_args)
                print("Snowflake connection established successfully.")
            except Exception as e:
                print(f"❌ Snowflake connection attempt failed: {str(e)}")
                raise e
                
            return self._conn

    # Reusing recommendation loading logic since this is from a static file for now
    def _text(self, val) -> str:
        if val is None:
            return ""
        return str(val).strip()

    def _parse_vaps_header(self, value: object) -> Tuple[str, str]:
        header = " ".join(str(value or "").replace("\r", "\n").split())
        match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", header)
        if not match:
            return header, ""
        return match.group(1).strip(), match.group(2).strip()

    def _recommendation_kind(self, value: object) -> str:
        value_text = self._text(value)
        lowered = value_text.lower()
        if not value_text or value_text == "0" or lowered in {"not applicable", "n/a"}:
            return "Not covered"
        if re.fullmatch(r"\d+(\.\d+)?", value_text):
            return "Fixed quantity"
        if value_text.startswith("# based on"):
            return "Quantity rule"
        if lowered.startswith("if ") or " if " in lowered:
            return "Conditional rule"
        if lowered.startswith("same as"):
            return "Dependency rule"
        return "Rule"

    def _load_recommendations(self) -> Dict[Tuple[str, str], RecommendationEntry]:
        path = os.path.join(settings.DATA_PATH, "actual_recomendationsheet.xlsx")
        if not os.path.exists(path):
            return {}
        
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        entries = {}

        header_row = None
        unit_col = None
        for row in ws.iter_rows():
            for cell in row:
                if self._text(cell.value) == "Unit - Product Code (SF)":
                    header_row = cell.row
                    unit_col = cell.column
                    break
            if header_row is not None:
                break
        
        if header_row is None:
            return {}

        vaps_columns = []
        for cell in ws[header_row]:
            vaps_desc, vaps_code = self._parse_vaps_header(cell.value)
            if cell.column >= 27 and re.fullmatch(r"[A-Z0-9]+", vaps_code):
                vaps_columns.append((cell.column, vaps_code, vaps_desc, len(vaps_columns)))

        for row_index in range(header_row + 1, ws.max_row + 1):
            unit_code = self._text(ws.cell(row_index, unit_col).value)
            if not unit_code:
                continue
            
            for column, vaps_code, vaps_desc, sequence in vaps_columns:
                val = ws.cell(row_index, column).value
                val_text = self._text(val)
                covered = val_text.lower() not in {"not applicable", "n/a"} and val_text != "0" and bool(val_text)
                
                entries[(unit_code, vaps_code)] = RecommendationEntry(
                    unit=unit_code,
                    vaps=vaps_code,
                    vapsDesc=vaps_desc,
                    recommendationValue=val_text,
                    coveredByRecommendationLogic=covered,
                    recommendationKind=self._recommendation_kind(val),
                    sequence=sequence
                )
        return entries

    def _execute_query(self, query: str, segment_key: Optional[str] = None, segment_col: Optional[str] = None) -> List[VapsAttachRate]:
        rows = []
        try:
            conn = self._get_connection()
            with conn.cursor(snowflake.connector.DictCursor) as cur:
                print(f"Executing Snowflake Query: {query[:150]}...")
                cur.execute(query)
                for row in cur:
                    vaps_code = self._text(row.get("vaps_code") or row.get("VAPS_CODE"))
                    unit_code = self._text(row.get("unit_code") or row.get("UNIT_CODE"))
                    if not vaps_code or not unit_code:
                        continue

                    rec = self.recommendations.get((unit_code, vaps_code))

                    segment_val = ""
                    if segment_col:
                        # Handle multiple possible column names (e.g. DIVISION or DIVISION_NAME)
                        if isinstance(segment_col, tuple):
                            for col in segment_col:
                                val = row.get(col) or row.get(col.upper())
                                if val:
                                    segment_val = self._text(val)
                                    break
                        else:
                            segment_val = self._text(row.get(segment_col) or row.get(segment_col.upper()))

                    attach_rate_val = row.get("Vaps_Attach_Rate") or row.get("VAPS_ATTACH_RATE")
                    attach_rate = float(attach_rate_val) / 100.0 if attach_rate_val is not None else 0.0

                    rows.append(VapsAttachRate(
                        unit=unit_code,
                        vaps=vaps_code,
                        vapsDesc=self._text(row.get("VAPS_DESCRIPTION")),
                        activations=int(row.get("Unit_Activations") or row.get("UNIT_ACTIVATIONS") or 0),
                        associated=int(row.get("Vaps_Associated_With_Unit") or row.get("VAPS_ASSOCIATED_WITH_UNIT") or 0),
                        attachRate=attach_rate,
                        unitName=self._text(row.get("UNIT_PRODUCTNAME_SF")),
                        unitDescription=self._text(row.get("UNIT_DESCRIPTION")),
                        unitL2=self._text(row.get("UNIT_L2_CORE_SOLUTION")),
                        unitL3=self._text(row.get("UNIT_L3_PRODUCTS")),
                        mainGroup=self._text(row.get("VAPS_MAIN_GROUP")) or "Unmapped",
                        detailedGroup=self._text(row.get("VAPS_DETAILED_GROUP")) or "Unmapped",
                        tier=self._text(row.get("VAPS_PACKAGE_TIER")) or "Unmapped",
                        source=self._text(row.get("VAPS_SOURCE")) or "Unmapped",
                        coveredByRecommendationLogic=rec.coveredByRecommendationLogic if rec else False,
                        market=segment_val if segment_key == "market" else "",
                        division=segment_val if segment_key == "division" else "",
                        region=segment_val if segment_key == "region" else ""
                    ))
        except Exception as e:
            print(f"❌ SNOWFLAKE ERROR in _execute_query: {str(e)}")
            raise e
        return rows

    @lru_cache(maxsize=1)
    def get_unit_attach_rates(self) -> List[VapsAttachRate]:
        query = f"SELECT * FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate"
        return self._execute_query(query)

    @lru_cache(maxsize=1)
    def get_market_attach_rates(self) -> List[VapsAttachRate]:
        query = f"SELECT * FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.GS_UNIT_MARKET_SEGMENT_VAPS_ATTACHED_RATE_NEW"
        return self._execute_query(query, "market", "MARKET_SEGMENT_DESCRIPTION")

    @lru_cache(maxsize=1)
    def get_division_attach_rates(self) -> List[VapsAttachRate]:
        query = f"SELECT * FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_division"
        return self._execute_query(query, "division", ("division", "DIVISION", "DIVISION_NAME"))

    @lru_cache(maxsize=1)
    def get_region_attach_rates(self) -> List[VapsAttachRate]:
        query = f"SELECT * FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_region"
        return self._execute_query(query, "region", ("region", "REGION", "REGION_DESCRIPTION"))

    def get_all_segments_data(self) -> Dict[str, List[VapsAttachRate]]:
        # For the segments tab, we need the full data. 
        # We fetch each one safely so one missing view doesn't crash the dashboard.
        segments = {"Market": [], "Division": [], "Region": []}
        
        try:
            segments["Market"] = self.get_market_attach_rates()
        except Exception as e:
            print(f"⚠️ Market data skipped: {str(e)}")

        try:
            segments["Division"] = self.get_division_attach_rates()
        except Exception as e:
            print(f"⚠️ Division data skipped: {str(e)}")

        try:
            segments["Region"] = self.get_region_attach_rates()
        except Exception as e:
            print(f"⚠️ Region data skipped: {str(e)}")

        return segments

    @lru_cache(maxsize=1)
    def get_metadata(self) -> Dict:
        """
        Optimized metadata fetch using DISTINCT queries.
        This is MUCH faster than pulling all rows.
        """
        print("Fetching optimized metadata from Snowflake...")
        metadata = {
            "sources": [],
            "groups": [],
            "markets": [],
            "divisions": [],
            "regions": []
        }
        
        queries = {
            "sources": [(f"SELECT DISTINCT VAPS_SOURCE FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate WHERE VAPS_SOURCE IS NOT NULL",)],
            "groups": [(f"SELECT DISTINCT VAPS_MAIN_GROUP FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate WHERE VAPS_MAIN_GROUP IS NOT NULL",)],
            "markets": [
                (f"SELECT DISTINCT MARKET_SEGMENT_DESCRIPTION FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.GS_UNIT_MARKET_SEGMENT_VAPS_ATTACHED_RATE_NEW WHERE MARKET_SEGMENT_DESCRIPTION IS NOT NULL",),
                (f"SELECT DISTINCT MARKET FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.GS_UNIT_MARKET_SEGMENT_VAPS_ATTACHED_RATE_NEW WHERE MARKET IS NOT NULL",)
            ],
            "divisions": [
                (f'SELECT DISTINCT "division" FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_division WHERE "division" IS NOT NULL',),
                (f"SELECT DISTINCT DIVISION FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_division WHERE DIVISION IS NOT NULL",),
                (f"SELECT DISTINCT DIVISION_NAME FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_division WHERE DIVISION_NAME IS NOT NULL",)
            ],
            "regions": [
                (f'SELECT DISTINCT "region" FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_region WHERE "region" IS NOT NULL',),
                (f"SELECT DISTINCT REGION FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_region WHERE REGION IS NOT NULL",),
                (f"SELECT DISTINCT REGION_DESCRIPTION FROM {settings.SNOWFLAKE_DATABASE}.{settings.SNOWFLAKE_SCHEMA}.gs_unit_vaps_attach_rate_region WHERE REGION_DESCRIPTION IS NOT NULL",)
            ]
        }

        try:
            conn = self._get_connection()
            with conn.cursor() as cur:
                for key, sql_list in queries.items():
                    success = False
                    for sql_tuple in sql_list:
                        sql = sql_tuple[0]
                        try:
                            print(f"Fetching Snowflake Metadata for {key} with query: {sql[:50]}...")
                            cur.execute(sql)
                            metadata[key] = sorted([str(r[0]) for r in cur.fetchall() if r[0]])
                            success = True
                            break # It worked, stop trying fallbacks
                        except Exception as e:
                            print(f"⚠️ Query failed for {key}, trying fallback if available: {str(e)}")
                    
                    if not success:
                        print(f"❌ All metadata queries failed for {key}.")
                        metadata[key] = []
        except Exception as e:
            print(f"❌ Critical error in get_metadata: {str(e)}")
            # Don't raise here, return whatever metadata we have
        
        return metadata

    def get_recommendation_entries(self) -> Dict[Tuple[str, str], RecommendationEntry]:
        return self.recommendations
