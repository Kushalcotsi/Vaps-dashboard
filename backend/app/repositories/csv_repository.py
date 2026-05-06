import pandas as pd
import os
import re
import logging
from typing import List, Dict, Tuple, Optional, Any
from app.repositories.base import BaseRepository
from app.models.dashboard import VapsAttachRate, RecommendationEntry
from app.core.config import settings
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class CSVRepository(BaseRepository):
    # Class-level cache to ensure data is loaded and PROCESSED only once
    _cache: Dict[str, Any] = {}
    _is_loaded = False

    def __init__(self):
        self.data_dir = settings.DATA_PATH
        if not CSVRepository._is_loaded:
            self._initialize_cache()

    def _initialize_cache(self):
        logger.info("Initializing Full In-Memory Processing Cache...")
        recommendations = self._load_recommendations()
        CSVRepository._cache["recommendations"] = recommendations
        
        # Files and their segment keys
        segment_configs = {
            "Market": ("unit_market_segment_vaps_attach_rate.csv", "market"),
            "Division": ("attach_rate_unit_division.csv", "division"),
            "Region": ("attach_rate_unit_region.csv", "region"),
            "Unit": ("unit_segment_vaps_attach_rate.csv", None)
        }
        
        for name, (filename, key) in segment_configs.items():
            path = os.path.join(self.data_dir, filename)
            if os.path.exists(path):
                logger.info(f"Loading and Processing {filename}...")
                df = pd.read_csv(path)
                # PROCESS ONCE and cache the list of objects
                CSVRepository._cache[f"obj_{name}"] = self._process_dataframe_to_objects(df, recommendations, segment_key=key)
            else:
                logger.warning(f"Data file {filename} not found!")
                CSVRepository._cache[f"obj_{name}"] = []
        
        CSVRepository._is_loaded = True
        logger.info("Full In-Memory Cache initialization complete.")

    def _text(self, val) -> str:
        if pd.isna(val):
            return ""
        return str(val or "").strip()

    def _number(self, val) -> float:
        if pd.isna(val):
            return 0.0
        try:
            return float(str(val or "").replace(",", ""))
        except ValueError:
            return 0.0

    def _parse_vaps_header(self, value: object) -> Tuple[str, str]:
        header = " ".join(str(value or "").replace("\r", "\n").split())
        match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", header)
        if not match:
            return header, ""
        return match.group(1).strip(), match.group(2).strip()

    def _recommendation_kind(self, value: object) -> str:
        value_text = self._text(value)
        lowered = value_text.lower()
        if not value_text or value_text == "0" or lowered in {"not applicable", "n/a"}:
            return "Not covered"
        if re.fullmatch(r"\d+(\.\d+)?", value_text):
            return "Fixed quantity"
        if value_text.startswith("# based on"):
            return "Quantity rule"
        if lowered.startswith("if ") or " if " in lowered:
            return "Conditional rule"
        if lowered.startswith("same as"):
            return "Dependency rule"
        return "Rule"

    def _load_recommendations(self) -> Dict[Tuple[str, str], RecommendationEntry]:
        path = os.path.join(self.data_dir, "actual_recomendationsheet.xlsx")
        if not os.path.exists(path):
            return {}
        
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        entries = {}

        header_row = None
        unit_col = None
        for row in ws.iter_rows():
            for cell in row:
                if self._text(cell.value) == "Unit - Product Code (SF)":
                    header_row = cell.row
                    unit_col = cell.column
                    break
            if header_row is not None:
                break
        
        if header_row is None:
            return {}

        vaps_columns = []
        for cell in ws[header_row]:
            vaps_desc, vaps_code = self._parse_vaps_header(cell.value)
            if cell.column >= 27 and re.fullmatch(r"[A-Z0-9]+", vaps_code):
                vaps_columns.append((cell.column, vaps_code, vaps_desc, len(vaps_columns)))

        for row_index in range(header_row + 1, ws.max_row + 1):
            unit_code = self._text(ws.cell(row_index, unit_col).value)
            if not unit_code:
                continue
            
            for column, vaps_code, vaps_desc, sequence in vaps_columns:
                val = ws.cell(row_index, column).value
                val_text = self._text(val)
                covered = val_text.lower() not in {"not applicable", "n/a"} and val_text != "0" and bool(val_text)
                
                entries[(unit_code, vaps_code)] = RecommendationEntry(
                    unit=unit_code,
                    vaps=vaps_code,
                    vapsDesc=vaps_desc,
                    recommendationValue=val_text,
                    coveredByRecommendationLogic=covered,
                    recommendationKind=self._recommendation_kind(val),
                    sequence=sequence
                )
        return entries

    def _process_dataframe_to_objects(self, df: pd.DataFrame, recommendations: Dict, segment_key: Optional[str] = None) -> List[VapsAttachRate]:
        """
        Private method to convert raw DF rows to Pydantic objects ONCE.
        """
        rows = []
        for _, source_row in df.iterrows():
            vaps_code = self._text(source_row.get("vaps_item_id") or source_row.get("vaps_code"))
            unit_code = self._text(source_row.get("UNIT_PRODUCTCODE_SF") or source_row.get("unit_code"))
            
            if not vaps_code or not unit_code:
                continue
                
            rec = recommendations.get((unit_code, vaps_code))
            
            market = ""
            division = ""
            region = ""
            
            if segment_key == "market":
                market = self._text(source_row.get("MARKET_SEGMENT_DESCRIPTION") or source_row.get("market"))
            elif segment_key == "division":
                division = self._text(source_row.get("division"))
            elif segment_key == "region":
                region = self._text(source_row.get("region"))
                division = self._text(source_row.get("division"))

            rows.append(VapsAttachRate(
                unit=unit_code,
                vaps=vaps_code,
                vapsDesc=self._text(source_row.get("VAPS_DESCRIPTION") or source_row.get("vaps_desc") or (rec.vapsDesc if rec else "")),
                activations=int(self._number(source_row.get("Unit_Activations") or source_row.get("unit_baskets_2024_26") or source_row.get("unit_baskets_2024"))),
                associated=int(self._number(source_row.get("Vaps_Associated_With_Unit") or source_row.get("vaps_baskets_2024_26") or source_row.get("vaps_baskets_2024"))),
                attachRate=self._number(source_row.get("Vaps_Attach_Rate") or source_row.get("attach_rate_2024_26") or source_row.get("attach_rate_2024")) / (100 if "Vaps_Attach_Rate" in source_row else 1),
                unitName=self._text(source_row.get("UNIT_PRODUCTNAME_SF") or (rec.unitName if rec else "")),
                unitDescription=self._text(source_row.get("UNIT_DESCRIPTION") or (rec.unitDescription if rec else "")),
                unitL2=self._text(source_row.get("UNIT_L2_CORE_SOLUTION") or (rec.unitL2 if rec else "")),
                unitL3=self._text(source_row.get("UNIT_L3_PRODUCTS") or (rec.unitL3 if rec else "")),
                mainGroup=self._text(source_row.get("VAPS_MAIN_GROUP") or source_row.get("vaps_main_group")) or "Unmapped",
                detailedGroup=self._text(source_row.get("VAPS_DETAILED_GROUP") or source_row.get("vaps_detailed_group")) or "Unmapped",
                tier=self._text(source_row.get("VAPS_PACKAGE_TIER") or source_row.get("vaps_package_tier")) or "Unmapped",
                source=self._text(source_row.get("VAPS_SOURCE") or source_row.get("vaps_source")) or "Unmapped",
                coveredByRecommendationLogic=rec.coveredByRecommendationLogic if rec else False,
                market=market,
                division=division,
                region=region
            ))
        return rows

    def get_all_segments_data(self) -> Dict[str, List[VapsAttachRate]]:
        # Return pre-processed objects directly
        return {
            "Market": CSVRepository._cache["obj_Market"],
            "Division": CSVRepository._cache["obj_Division"],
            "Region": CSVRepository._cache["obj_Region"]
        }

    def get_unit_attach_rates(self) -> List[VapsAttachRate]:
        return CSVRepository._cache["obj_Unit"]

    def get_market_attach_rates(self) -> List[VapsAttachRate]:
        return CSVRepository._cache["obj_Market"]

    def get_division_attach_rates(self) -> List[VapsAttachRate]:
        return CSVRepository._cache["obj_Division"]

    def get_region_attach_rates(self) -> List[VapsAttachRate]:
        return CSVRepository._cache["obj_Region"]

    def get_recommendation_entries(self) -> Dict[Tuple[str, str], RecommendationEntry]:
        return CSVRepository._cache["recommendations"]
